"""
modules/case_management.py
══════════════════════════════════════════════════════════════════════════
Case Management System — Enhanced with Dynamic Duplicate Studio

Features:
  • Dynamic Duplicate Studio (single-col, multi-col, fuzzy detection)
  • Auto column profiling & key recommendations (Strong / Medium / Weak)
  • Fuzzy duplicate detection with similarity threshold slider
  • Survivorship rule selector (Most Complete, Most Recent, Source Priority…)
  • Automatic case creation per duplicate group
  • Golden record generation with group-by-group comparison
  • Duplicate Analytics Dashboard
  • Case CRUD with status workflow & audit trail
  • Excel report export (multi-sheet)
══════════════════════════════════════════════════════════════════════════
"""

import datetime
import uuid
from io import BytesIO
from collections import defaultdict
from difflib import SequenceMatcher
from typing import Dict, List, Tuple, Any, Optional

import streamlit as st
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from matplotlib.patches import Wedge, Rectangle

from modules.config import AppConfig
from modules.ui_components import UIComponents


# ══════════════════════════════════════════════════════════════════════════
#  CONSTANTS & STYLING
# ══════════════════════════════════════════════════════════════════════════

_CASE_STATUSES   = ["Open", "In Progress", "Under Review", "Resolved", "Closed"]
_CASE_PRIORITIES = ["Critical", "High", "Medium", "Low"]
_CASE_TYPES = [
    "Duplicate Records", "Missing Values", "Invalid Format",
    "Outlier / Range Violation", "Uniqueness Violation",
    "Standardization Issue", "Consistency Issue", "Other",
]
_SURVIVORSHIP_RULES = [
    "Most Complete", "Most Recent", "Most Frequent",
    "Source Priority", "Manual Selection",
]

STATUS_COLORS = {
    "Open": "#ef4444", "In Progress": "#f59e0b",
    "Under Review": "#3b82f6", "Resolved": "#10b981", "Closed": "#6b7280",
}
PRIORITY_COLORS = {
    "Critical": "#dc2626", "High": "#f97316",
    "Medium": "#eab308", "Low": "#22c55e",
}

_GDG_LIGHT_STYLE = """
<style>
:root,[data-testid="stDataEditor"],[data-testid="stDataEditor"] > div {
    --gdg-bg-cell:#ffffff !important;--gdg-bg-cell-medium:#f7f4fc !important;
    --gdg-bg-header:#ede8f7 !important;--gdg-bg-header-has-focus:#e0d9f2 !important;
    --gdg-bg-header-hovered:#d4cced !important;--gdg-border-color:#e8e2f5 !important;
    --gdg-horizontal-border-color:#e8e2f5 !important;--gdg-accent-color:#7c3aed !important;
    --gdg-accent-light:rgba(124,58,237,0.10) !important;--gdg-text-dark:#1a1028 !important;
    --gdg-text-medium:#3b2f54 !important;--gdg-text-light:#6b5f82 !important;
    --gdg-text-header:#3b1f72 !important;--gdg-text-header-selected:#1a0a40 !important;
    --gdg-cell-text-color:#1a1028 !important;
}
[data-testid="stDataEditor"] canvas{background-color:#ffffff !important;}
[data-testid="stDataEditor"] .dvn-scroller,[data-testid="stDataEditor"] .dvn-scroll-inner,
[data-testid="stDataEditor"] > div,[data-testid="stDataEditor"] > div > div{background:#ffffff !important;}
</style>
"""


# ══════════════════════════════════════════════════════════════════════════
#  SESSION STATE INITIALIZATION
# ══════════════════════════════════════════════════════════════════════════

def init_case_management_state() -> None:
    """Initialize all session-state keys for Case Management."""
    defaults = {
        "cases":               [],
        "case_counter":        0,
        "case_excel_reports":  {},
        "dup_groups":          None,
        "dup_golden_records":  {},
        "dup_source_df":       None,
        "dup_match_columns":   [],
        "case_filter_status":  "All",
        "case_filter_priority":"All",
        "case_filter_type":    "All",
        # Studio-specific
        "studio_dup_df":       None,
        "studio_profile":      None,
        "studio_fuzzy_groups": None,
        "studio_match_mode":   "Exact (Single Column)",
    }
    for k, v in defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v


# ══════════════════════════════════════════════════════════════════════════
#  CASE CRUD HELPERS
# ══════════════════════════════════════════════════════════════════════════

def _next_case_id() -> str:
    st.session_state["case_counter"] += 1
    return f"CASE-{st.session_state['case_counter']:04d}"


def create_case(
    title: str,
    case_type: str,
    priority: str = "Medium",
    description: str = "",
    affected_records: int = 0,
    affected_columns: str = "",
    source: str = "Manual",
    extra: Optional[dict] = None,
) -> dict:
    """Create a new case and append to session state."""
    now = datetime.datetime.now()
    case = {
        "case_id":          _next_case_id(),
        "title":            title,
        "type":             case_type,
        "priority":         priority,
        "status":           "Open",
        "description":      description,
        "affected_records": affected_records,
        "affected_columns": affected_columns,
        "source":           source,
        "created_at":       now.strftime("%Y-%m-%d %H:%M"),
        "updated_at":       now.strftime("%Y-%m-%d %H:%M"),
        "resolved_at":      "",
        "assigned_to":      "",
        "history":          [{"ts": now.strftime("%Y-%m-%d %H:%M"),
                              "action": "Case created", "by": "System"}],
        "extra":            extra or {},
    }
    st.session_state["cases"].append(case)
    return case


def update_case_status(case_id: str, new_status: str, note: str = "", by: str = "User") -> None:
    now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
    for c in st.session_state["cases"]:
        if c["case_id"] == case_id:
            old = c["status"]
            c["status"]     = new_status
            c["updated_at"] = now
            if new_status in ("Resolved", "Closed"):
                c["resolved_at"] = now
            c["history"].append({
                "ts": now,
                "action": f"Status changed: {old} → {new_status}" + (f" ({note})" if note else ""),
                "by": by,
            })
            break


def auto_create_cases_from_dq(results_df: pd.DataFrame, dim_scores: dict) -> int:
    if results_df is None or results_df.empty:
        return 0
    created = 0
    existing = {c["title"] for c in st.session_state["cases"]}
    for dim, score in (dim_scores or {}).items():
        if score < 80:
            title = f"DQ Issue: {dim} score {score:.1f}%"
            if title not in existing:
                prio = "Critical" if score < 50 else ("High" if score < 70 else "Medium")
                failed = results_df[results_df["Issue categories"].str.contains(dim, na=False)]
                create_case(
                    title=title,
                    case_type=_map_dim_to_case_type(dim),
                    priority=prio,
                    description=f"Dimension '{dim}' scored {score:.1f}% — below 80% threshold.",
                    affected_records=len(failed),
                    source="DQ Engine",
                )
                created += 1
    dup_rows = results_df[results_df["Failed_Rules"].str.contains("uniqueness", na=False, case=False)]
    if len(dup_rows) > 0:
        title = f"Duplicate Records Detected ({len(dup_rows)} rows)"
        if title not in existing:
            create_case(
                title=title, case_type="Duplicate Records", priority="High",
                description=f"{len(dup_rows)} records flagged for uniqueness violations.",
                affected_records=len(dup_rows), source="DQ Engine",
            )
            created += 1
    return created


def _map_dim_to_case_type(dim: str) -> str:
    return {
        "Completeness": "Missing Values", "Validity": "Invalid Format",
        "Uniqueness": "Uniqueness Violation", "Standardization": "Standardization Issue",
        "Consistency": "Consistency Issue", "Accuracy": "Outlier / Range Violation",
    }.get(dim, "Other")


# ══════════════════════════════════════════════════════════════════════════
#  COLUMN PROFILER — auto-recommendation engine
# ══════════════════════════════════════════════════════════════════════════

def profile_columns(df: pd.DataFrame) -> pd.DataFrame:
    """
    Profile every column for uniqueness %, null %, cardinality.
    Returns a DataFrame with recommendation: Strong / Medium / Weak identifier.
    """
    records = []
    total = len(df)
    non_internal = [c for c in df.columns if not c.startswith("_")]

    for col in non_internal:
        series = df[col].astype(str).str.strip().str.lower()
        null_pct        = df[col].isna().mean() * 100
        empty_pct       = (series == "").mean() * 100 + (series == "nan").mean() * 100
        effective_null  = null_pct + empty_pct
        cardinality     = series[series.notna() & (series != "") & (series != "nan")].nunique()
        uniqueness_pct  = (cardinality / total * 100) if total > 0 else 0

        if uniqueness_pct >= 90 and effective_null < 5:
            strength = "🟢 Strong"
        elif uniqueness_pct >= 50 and effective_null < 20:
            strength = "🟡 Medium"
        else:
            strength = "🔴 Weak"

        records.append({
            "Column":         col,
            "Cardinality":    cardinality,
            "Uniqueness %":   round(uniqueness_pct, 1),
            "Null/Empty %":   round(effective_null, 1),
            "Recommendation": strength,
        })

    return pd.DataFrame(records).sort_values("Uniqueness %", ascending=False)


# ══════════════════════════════════════════════════════════════════════════
#  DUPLICATE DETECTION ENGINE
# ══════════════════════════════════════════════════════════════════════════

def detect_duplicates(
    df: pd.DataFrame,
    match_columns: List[str],
    fuzzy: bool = False,
    threshold: float = 0.85,
) -> pd.DataFrame:
    """
    Detect duplicate groups. Returns annotated DataFrame with:
      _dup_group_id, _is_duplicate, _dup_count,
      _completeness, _recency_rank, _match_type, _similarity_score
    """
    result = df.copy()
    result["_dup_group_id"]      = np.nan
    result["_is_duplicate"]      = False
    result["_dup_count"]         = 0
    result["_match_type"]        = ""
    result["_similarity_score"]  = 0.0

    non_internal = [c for c in df.columns if not c.startswith("_")]
    result["_completeness"] = (
        df[non_internal].notna().sum(axis=1) / max(len(non_internal), 1) * 100
    ).round(2)

    # Normalize for matching
    match_df = df[match_columns].copy()
    for col in match_columns:
        match_df[col] = match_df[col].astype(str).str.strip().str.lower().fillna("")

    match_df["_match_key"] = match_df[match_columns].apply(
        lambda r: "|".join(r.values.astype(str)), axis=1
    )

    group_id = 0

    # ── Exact matching ─────────────────────────────────────────────────────
    key_counts = match_df["_match_key"].value_counts()
    dup_keys   = key_counts[key_counts > 1].index

    for key in dup_keys:
        idxs = match_df[match_df["_match_key"] == key].index.tolist()
        group_id += 1
        gid = f"DG-{group_id:04d}"
        for idx in idxs:
            result.at[idx, "_dup_group_id"]     = gid
            result.at[idx, "_is_duplicate"]     = True
            result.at[idx, "_dup_count"]        = len(idxs)
            result.at[idx, "_match_type"]       = "Exact"
            result.at[idx, "_similarity_score"] = 1.0

    # ── Fuzzy matching ─────────────────────────────────────────────────────
    if fuzzy and len(match_columns) == 1:
        col = match_columns[0]
        keys = match_df["_match_key"].tolist()
        indices = match_df.index.tolist()
        unassigned = [
            (i, idx) for i, idx in enumerate(indices)
            if not result.at[idx, "_is_duplicate"]
        ]

        visited: set = set()
        fuzzy_groups: List[List[int]] = []

        for i, idx_a in unassigned:
            if idx_a in visited:
                continue
            grp = [idx_a]
            for j, idx_b in unassigned:
                if idx_b == idx_a or idx_b in visited:
                    continue
                sim = SequenceMatcher(None, keys[i], keys[j]).ratio()
                if sim >= threshold:
                    grp.append(idx_b)
                    result.at[idx_b, "_similarity_score"] = round(sim, 3)
            if len(grp) > 1:
                fuzzy_groups.append(grp)
                visited.update(grp)

        for grp in fuzzy_groups:
            group_id += 1
            gid = f"DG-{group_id:04d}-F"
            for idx in grp:
                result.at[idx, "_dup_group_id"] = gid
                result.at[idx, "_is_duplicate"] = True
                result.at[idx, "_dup_count"]    = len(grp)
                result.at[idx, "_match_type"]   = "Fuzzy"
                if result.at[idx, "_similarity_score"] == 0.0:
                    result.at[idx, "_similarity_score"] = 1.0

        st.session_state["studio_fuzzy_groups"] = fuzzy_groups

    # ── Recency rank ───────────────────────────────────────────────────────
    date_cols = df.select_dtypes(include=["datetime", "datetime64"]).columns.tolist()
    if date_cols:
        result["_recency_rank"] = (
            result.groupby("_dup_group_id")[date_cols[0]]
            .rank(method="first", ascending=False).fillna(0).astype(int)
        )
    else:
        result["_recency_rank"] = (
            result.groupby("_dup_group_id").cumcount(ascending=False) + 1
        )

    return result


def _auto_create_cases_for_dup_groups(
    dup_df: pd.DataFrame,
    match_columns: List[str],
    match_type: str = "Exact",
) -> int:
    """
    For each duplicate group detected, create one case automatically.
    Returns count of new cases created.
    """
    created = 0
    existing = {c["title"] for c in st.session_state["cases"]}
    dup_only = dup_df[dup_df["_is_duplicate"]]

    for gid, grp in dup_only.groupby("_dup_group_id"):
        title = f"Dup Group {gid}: {len(grp)} records on [{', '.join(match_columns)}]"
        if title not in existing:
            row_indices = grp.index.tolist()
            sim_scores  = grp["_similarity_score"].tolist() if "_similarity_score" in grp.columns else []
            create_case(
                title=title,
                case_type="Duplicate Records",
                priority="High",
                description=(
                    f"Duplicate group {gid} contains {len(grp)} records matched "
                    f"via {match_type} comparison on columns: {', '.join(match_columns)}."
                    + (f" Avg similarity: {sum(sim_scores)/len(sim_scores):.2%}" if sim_scores else "")
                ),
                affected_records=len(grp),
                affected_columns=", ".join(match_columns),
                source="Dynamic Duplicate Studio",
                extra={
                    "group_id":      gid,
                    "match_type":    match_type,
                    "row_indices":   row_indices,
                    "match_columns": match_columns,
                    "record_count":  len(grp),
                },
            )
            created += 1
    return created


# ══════════════════════════════════════════════════════════════════════════
#  GOLDEN RECORD GENERATION
# ══════════════════════════════════════════════════════════════════════════

def identify_golden_record(group_df: pd.DataFrame, strategy: str = "Most Complete") -> int:
    if group_df.empty:
        return -1
    if strategy == "Most Complete":
        return int(group_df["_completeness"].idxmax())
    elif strategy == "Most Recent":
        return int(group_df["_recency_rank"].idxmin())
    elif strategy == "Most Frequent":
        non_internal = [c for c in group_df.columns if not c.startswith("_")]
        return int(group_df[non_internal].notna().sum(axis=1).idxmax())
    elif strategy == "Source Priority":
        # Find any column named 'source', 'system', 'origin' etc.
        src_cols = [c for c in group_df.columns
                    if any(kw in c.lower() for kw in ["source", "system", "origin", "priority"])]
        if src_cols:
            return int(group_df[src_cols[0]].idxmin())  # alphabetically first = highest priority
        return int(group_df["_completeness"].idxmax())
    else:
        return int(group_df.index[0])


def build_golden_records_df(
    dup_df: pd.DataFrame,
    strategy: str = "Most Complete",
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """Build (golden_df, discards_df) from annotated duplicate dataframe."""
    non_dup  = dup_df[~dup_df["_is_duplicate"]].copy()
    dup_only = dup_df[dup_df["_is_duplicate"]].copy()

    golden_rows: List[pd.DataFrame] = []
    discard_rows: List[pd.DataFrame] = []

    for gid, grp in dup_only.groupby("_dup_group_id"):
        golden_idx = identify_golden_record(grp, strategy)
        if golden_idx == -1:
            continue
        golden_rows.append(grp.loc[[golden_idx]])
        discard_rows.append(grp.drop(index=golden_idx))

    golden_df   = pd.concat([non_dup] + golden_rows,  ignore_index=False) if golden_rows  else non_dup.copy()
    discards_df = pd.concat(discard_rows, ignore_index=False)              if discard_rows else pd.DataFrame()

    return golden_df, discards_df


# ══════════════════════════════════════════════════════════════════════════
#  EXCEL REPORT BUILDER
# ══════════════════════════════════════════════════════════════════════════

def build_case_excel(
    cases: List[dict],
    dup_df: Optional[pd.DataFrame] = None,
    golden_df: Optional[pd.DataFrame] = None,
    discards_df: Optional[pd.DataFrame] = None,
) -> bytes:
    """Build a multi-sheet Excel workbook with all case management data."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    header_fill  = PatternFill(start_color="6d28d9", end_color="7c3aed", fill_type="solid")
    header_font  = Font(bold=True, color="FFFFFF", size=11)
    thin_border  = Border(
        left=Side(style="thin", color="d6d3d1"),  right=Side(style="thin", color="d6d3d1"),
        top=Side(style="thin", color="d6d3d1"),   bottom=Side(style="thin", color="d6d3d1"),
    )
    golden_fill  = PatternFill(start_color="DCFCE7", end_color="BBF7D0", fill_type="solid")
    discard_fill = PatternFill(start_color="FEE2E2", end_color="FECACA", fill_type="solid")

    def _write_df(ws, df: pd.DataFrame, start_row: int = 1, highlight_golden: bool = False):
        cols = [c for c in df.columns if not c.startswith("_") or c in ("_is_golden", "_dup_group_id", "_completeness", "_dup_count", "_similarity_score", "_match_type")]
        for ci, col in enumerate(cols, 1):
            cell = ws.cell(row=start_row, column=ci, value=col)
            cell.fill, cell.font = header_fill, header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
        for ri, (_, row) in enumerate(df.iterrows(), start_row + 1):
            for ci, col in enumerate(cols, 1):
                val = row.get(col, "")
                if isinstance(val, (list, dict, np.ndarray)):
                    val = str(val)
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                if highlight_golden and "_is_golden" in df.columns:
                    cell.fill = golden_fill if row.get("_is_golden") else discard_fill
        for ci, col in enumerate(cols, 1):
            ws.column_dimensions[ws.cell(row=1, column=ci).column_letter].width = max(14, len(col) + 4)

    # ── Sheet 1: Case Summary ─────────────────────────────────────────────
    ws_cases = wb.active
    ws_cases.title = "Case Summary"
    case_cols = ["case_id", "title", "type", "priority", "status",
                 "affected_records", "source", "created_at", "updated_at", "resolved_at"]
    for ci, col in enumerate(case_cols, 1):
        cell = ws_cases.cell(row=1, column=ci, value=col.replace("_", " ").title())
        cell.fill = header_fill; cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    for ri, case in enumerate(cases, 2):
        for ci, col in enumerate(case_cols, 1):
            cell = ws_cases.cell(row=ri, column=ci, value=str(case.get(col, "")))
            cell.border = thin_border
            if col == "status":
                sc = case.get("status", "")
                if sc == "Open":
                    cell.fill = PatternFill(start_color="FEE2E2", fill_type="solid")
                elif sc == "Resolved":
                    cell.fill = PatternFill(start_color="DCFCE7", fill_type="solid")
                elif sc == "In Progress":
                    cell.fill = PatternFill(start_color="FEF3C7", fill_type="solid")
    for ci, col in enumerate(case_cols, 1):
        ws_cases.column_dimensions[ws_cases.cell(row=1, column=ci).column_letter].width = max(16, len(col) + 6)

    # ── Sheet 2: Duplicate Groups ─────────────────────────────────────────
    if dup_df is not None and not dup_df.empty:
        dup_only = dup_df[dup_df["_is_duplicate"]].copy()
        if not dup_only.empty:
            ws_dup = wb.create_sheet("Duplicate Groups")
            _write_df(ws_dup, dup_only.sort_values("_dup_group_id"))

    # ── Sheet 3: Golden Records ───────────────────────────────────────────
    if golden_df is not None and not golden_df.empty:
        ws_gold = wb.create_sheet("Golden Records")
        display = [c for c in golden_df.columns if not c.startswith("_")]
        _write_df(ws_gold, golden_df[display])

    # ── Sheet 4: Discarded Records ────────────────────────────────────────
    if discards_df is not None and not discards_df.empty:
        ws_disc = wb.create_sheet("Discarded Records")
        display = [c for c in discards_df.columns if not c.startswith("_")]
        _write_df(ws_disc, discards_df[display])

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()


# ══════════════════════════════════════════════════════════════════════════
#  VISUALIZATION HELPERS
# ══════════════════════════════════════════════════════════════════════════

def _case_status_pie_png(cases: List[dict]) -> Optional[bytes]:
    if not cases:
        return None
    status_counts: Dict[str, int] = defaultdict(int)
    for c in cases:
        status_counts[c["status"]] += 1
    labels = list(status_counts.keys())
    sizes  = list(status_counts.values())
    colors = [STATUS_COLORS.get(l, "#94a3b8") for l in labels]
    fig, ax = plt.subplots(figsize=(4, 4), dpi=140)
    fig.patch.set_facecolor("#fafafa")
    wedges, texts, autotexts = ax.pie(
        sizes, labels=labels, colors=colors, autopct="%1.0f%%",
        startangle=90, pctdistance=0.78,
        wedgeprops=dict(width=0.42, edgecolor="white", linewidth=2.5),
        textprops=dict(fontsize=10, fontweight=600, color="#1c1917"),
    )
    for at in autotexts:
        at.set_fontsize(9); at.set_color("white"); at.set_fontweight("bold")
    ax.text(0, 0, f"{len(cases)}\nCases", ha="center", va="center",
            fontsize=16, fontweight="bold", color="#6d28d9")
    plt.tight_layout()
    buf = BytesIO()
    fig.savefig(buf, format="png", bbox_inches="tight", facecolor="#fafafa")
    plt.close(fig)
    return buf.getvalue()


def _case_priority_bar_png(cases: List[dict]) -> Optional[bytes]:
    if not cases:
        return None
    prio_counts: Dict[str, int] = defaultdict(int)
    for c in cases:
        prio_counts[c["priority"]] += 1
    ordered = [p for p in _CASE_PRIORITIES if p in prio_counts]
    sizes  = [prio_counts[p] for p in ordered]
    colors = [PRIORITY_COLORS.get(p, "#94a3b8") for p in ordered]
    fig, ax = plt.subplots(figsize=(5, 2.5), dpi=140)
    fig.patch.set_facecolor("#fafafa"); ax.set_facecolor("#ffffff")
    bars = ax.barh(ordered, sizes, color=colors, height=0.55, edgecolor="white", linewidth=2)
    ax.set_xlim(0, max(sizes) * 1.35 if sizes else 5)
    ax.set_xlabel("Count", color="#1c1917", fontsize=10, weight=600)
    ax.tick_params(colors="#44403c", labelsize=10)
    ax.spines[["top", "right", "bottom"]].set_visible(False)
    ax.spines["left"].set_color("#d6d3d1")
    for bar, cnt in zip(bars, sizes):
        ax.text(bar.get_width() + 0.2, bar.get_y() + bar.get_height() / 2,
                str(cnt), va="center", fontsize=11, fontweight="bold", color="#1c1917")
    plt.tight_layout()
    buf = BytesIO()
    fig.savefig(buf, format="png", bbox_inches="tight", facecolor="#fafafa")
    plt.close(fig)
    return buf.getvalue()


def _dup_group_bar_png(dup_df: pd.DataFrame) -> Optional[bytes]:
    if dup_df is None or dup_df.empty:
        return None
    groups = (
        dup_df[dup_df["_is_duplicate"]]
        .groupby("_dup_group_id").size()
        .sort_values(ascending=False).head(20)
    )
    if groups.empty:
        return None
    fig, ax = plt.subplots(figsize=(8, max(3, len(groups) * 0.5)), dpi=140)
    fig.patch.set_facecolor("#fafafa"); ax.set_facecolor("#ffffff")
    colors = ["#7c3aed" if v > 2 else "#a78bfa" for v in groups.values]
    bars = ax.barh(groups.index.astype(str), groups.values, color=colors,
                   height=0.6, edgecolor="white", linewidth=2)
    ax.set_xlabel("Records in Group", fontsize=10, weight=600, color="#1c1917")
    ax.set_title("Duplicate Groups (Top 20)", fontsize=12, weight=700, color="#6d28d9", pad=12)
    ax.tick_params(colors="#44403c", labelsize=9)
    ax.spines[["top", "right", "bottom"]].set_visible(False)
    ax.spines["left"].set_color("#d6d3d1")
    for bar, cnt in zip(bars, groups.values):
        ax.text(bar.get_width() + 0.15, bar.get_y() + bar.get_height() / 2,
                str(cnt), va="center", fontsize=10, fontweight="bold", color="#1c1917")
    plt.tight_layout()
    buf = BytesIO()
    fig.savefig(buf, format="png", bbox_inches="tight", facecolor="#fafafa")
    plt.close(fig)
    return buf.getvalue()


def _dup_analytics_charts_png(dup_df: pd.DataFrame) -> Dict[str, Optional[bytes]]:
    """Generate all duplicate analytics charts."""
    charts: Dict[str, Optional[bytes]] = {}

    if dup_df is None or dup_df.empty:
        return charts

    dup_only = dup_df[dup_df["_is_duplicate"]]

    # ── Match type distribution ───────────────────────────────────────────
    if "_match_type" in dup_only.columns and not dup_only.empty:
        type_counts = dup_only["_match_type"].value_counts()
        if not type_counts.empty:
            fig, ax = plt.subplots(figsize=(4, 3), dpi=130)
            fig.patch.set_facecolor("#fafafa")
            ax.bar(type_counts.index, type_counts.values,
                   color=["#7c3aed", "#a78bfa", "#c4b5fd"][:len(type_counts)],
                   edgecolor="white", linewidth=2)
            ax.set_title("Duplicate Type Distribution", fontsize=11, weight=700, color="#6d28d9")
            ax.set_ylabel("Records"); ax.spines[["top", "right"]].set_visible(False)
            plt.tight_layout()
            buf = BytesIO()
            fig.savefig(buf, format="png", bbox_inches="tight", facecolor="#fafafa")
            plt.close(fig)
            charts["match_type"] = buf.getvalue()

    # ── Fuzzy similarity distribution ─────────────────────────────────────
    if "_similarity_score" in dup_only.columns:
        fuzzy_only = dup_only[dup_only.get("_match_type", pd.Series(dtype=str)) == "Fuzzy"]
        if not fuzzy_only.empty and "_similarity_score" in fuzzy_only.columns:
            scores = fuzzy_only["_similarity_score"].dropna()
            if len(scores) > 0:
                fig, ax = plt.subplots(figsize=(5, 3), dpi=130)
                fig.patch.set_facecolor("#fafafa")
                ax.hist(scores, bins=10, color="#7c3aed", edgecolor="white", linewidth=1.5)
                ax.set_title("Fuzzy Similarity Distribution", fontsize=11, weight=700, color="#6d28d9")
                ax.set_xlabel("Similarity Score"); ax.set_ylabel("Count")
                ax.spines[["top", "right"]].set_visible(False)
                plt.tight_layout()
                buf = BytesIO()
                fig.savefig(buf, format="png", bbox_inches="tight", facecolor="#fafafa")
                plt.close(fig)
                charts["fuzzy_dist"] = buf.getvalue()

    return charts


def _golden_vs_discard_pie_png(golden_count: int, discard_count: int) -> Optional[bytes]:
    if golden_count == 0 and discard_count == 0:
        return None
    fig, ax = plt.subplots(figsize=(4, 4), dpi=140)
    fig.patch.set_facecolor("#fafafa")
    sizes  = [golden_count, discard_count]
    labels = ["Golden Records", "Discarded"]
    colors = ["#10b981", "#f87171"]
    wedges, texts, autotexts = ax.pie(
        sizes, labels=labels, colors=colors, autopct="%1.0f%%",
        startangle=90, pctdistance=0.78,
        wedgeprops=dict(width=0.42, edgecolor="white", linewidth=2.5),
        textprops=dict(fontsize=10, fontweight=600, color="#1c1917"),
    )
    for at in autotexts:
        at.set_fontsize(9); at.set_color("white"); at.set_fontweight("bold")
    total = golden_count + discard_count
    ax.text(0, 0, f"{total}\nTotal", ha="center", va="center",
            fontsize=15, fontweight="bold", color="#6d28d9")
    plt.tight_layout()
    buf = BytesIO()
    fig.savefig(buf, format="png", bbox_inches="tight", facecolor="#fafafa")
    plt.close(fig)
    return buf.getvalue()


# ══════════════════════════════════════════════════════════════════════════
#  MAIN PAGE RENDERER
# ══════════════════════════════════════════════════════════════════════════

def page_case_management():
    """Full Case Management page with tabs."""
    st.markdown(_GDG_LIGHT_STYLE, unsafe_allow_html=True)
    init_case_management_state()

    with st.sidebar:
        st.markdown("### 🧭 Navigation")
        if st.button("🏠 Home",     use_container_width=True, key="case_home"):
            st.session_state["page"] = "home"; st.rerun()
        if st.button("🔍 DQ",       use_container_width=True, key="case_dq"):
            st.session_state["page"] = "dq"; st.rerun()
        if st.button("📈 Maturity", use_container_width=True, key="case_maturity"):
            st.session_state["page"] = "maturity"; st.rerun()
        if st.button("📋 Policies", use_container_width=True, key="case_policy"):
            st.session_state["page"] = "policy"; st.rerun()
        st.divider()
        dq_results = st.session_state.get("dq_results_df")
        dq_dims    = st.session_state.get("dq_dim_scores")
        if dq_results is not None:
            st.markdown("### ⚡ Quick Actions")
            if st.button("🔄 Import DQ Issues as Cases", use_container_width=True, key="cm_auto"):
                n = auto_create_cases_from_dq(dq_results, dq_dims)
                st.success(f"✅ Created {n} new case(s)") if n else st.info("No new cases to create.")
                st.rerun()
            st.divider()

    st.markdown("# 🎯 Case Management System")
    st.markdown(
        "Track and resolve data quality issues, detect duplicates with the "
        "**Dynamic Duplicate Studio**, identify golden records, and maintain audit trails."
    )
    st.divider()

    tab_dash, tab_cases, tab_studio, tab_golden, tab_reports = st.tabs([
        "📊 Dashboard",
        "📋 Cases",
        "🔬 Dynamic Duplicate Studio",
        "🏆 Golden Records",
        "📥 Reports & Export",
    ])

    with tab_dash:
        _render_dashboard()
    with tab_cases:
        _render_cases_tab()
    with tab_studio:
        _render_dynamic_duplicate_studio()
    with tab_golden:
        _render_golden_records_tab()
    with tab_reports:
        _render_reports_tab()


# ══════════════════════════════════════════════════════════════════════════
#  TAB: DASHBOARD
# ══════════════════════════════════════════════════════════════════════════

def _render_dashboard():
    cases  = st.session_state["cases"]
    dup_df = st.session_state.get("dup_groups")

    if not cases and dup_df is None:
        st.markdown("""
        <div style="text-align:center;padding:3rem 0;">
            <div style="font-size:4rem;margin-bottom:1rem;">📭</div>
            <h3 style="color:#6d28d9;">No Cases Yet</h3>
            <p style="color:#57534e;">
                Create a case in <strong>Cases</strong> tab, run the 
                <strong>Dynamic Duplicate Studio</strong>, or import from a DQ Assessment.
            </p>
        </div>
        """, unsafe_allow_html=True)
        return

    total       = len(cases)
    open_cnt    = sum(1 for c in cases if c["status"] == "Open")
    ip_cnt      = sum(1 for c in cases if c["status"] == "In Progress")
    resolved    = sum(1 for c in cases if c["status"] in ("Resolved", "Closed"))
    critical    = sum(1 for c in cases if c["priority"] == "Critical" and c["status"] not in ("Resolved", "Closed"))

    m1, m2, m3, m4, m5 = st.columns(5)
    m1.metric("Total Cases", total)
    m2.metric("Open", open_cnt)
    m3.metric("In Progress", ip_cnt)
    m4.metric("Resolved / Closed", resolved)
    m5.metric("Critical (Active)", critical)

    UIComponents.render_micro_progress(
        int(resolved / total * 100) if total else 0, "#10b981", "#34d399"
    )
    st.divider()

    c1, c2 = st.columns(2)
    with c1:
        st.markdown("#### Case Status Distribution")
        img = _case_status_pie_png(cases)
        if img:
            st.image(img, use_container_width=True)
    with c2:
        st.markdown("#### Cases by Priority")
        img = _case_priority_bar_png(cases)
        if img:
            st.image(img, use_container_width=True)

    # ── Duplicate analytics summary ───────────────────────────────────────
    if dup_df is not None and not dup_df.empty:
        st.divider()
        st.markdown("#### 🔁 Duplicate Analytics Overview")
        dup_only = dup_df[dup_df["_is_duplicate"]]
        d1, d2, d3, d4 = st.columns(4)
        d1.metric("Total Records",     f"{len(dup_df):,}")
        d2.metric("Duplicate Records", f"{len(dup_only):,}")
        d3.metric("Duplicate Groups",  f"{dup_only['_dup_group_id'].nunique():,}")
        d4.metric("Unique Records",    f"{len(dup_df) - len(dup_only):,}")

        analytics = _dup_analytics_charts_png(dup_df)
        if analytics:
            acols = st.columns(len(analytics))
            for i, (key, img) in enumerate(analytics.items()):
                if img:
                    with acols[i]:
                        st.image(img, use_container_width=True)

    st.divider()
    st.markdown("#### Recent Cases")
    recent = sorted(cases, key=lambda c: c["created_at"], reverse=True)[:10]
    rows = [{
        "Case ID": c["case_id"], "Title": c["title"], "Type": c["type"],
        "Priority": c["priority"], "Status": c["status"],
        "Records": c["affected_records"], "Created": c["created_at"],
    } for c in recent]
    if rows:
        st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)


# ══════════════════════════════════════════════════════════════════════════
#  TAB: CASES (CRUD)
# ══════════════════════════════════════════════════════════════════════════

def _render_cases_tab():
    cases = st.session_state["cases"]

    with st.expander("➕ Create New Case", expanded=not bool(cases)):
        cc1, cc2 = st.columns(2)
        with cc1:
            new_title   = st.text_input("Case Title", placeholder="e.g. Duplicate vendor records", key="cm_title")
            new_type    = st.selectbox("Case Type", _CASE_TYPES, key="cm_type")
            new_prio    = st.selectbox("Priority", _CASE_PRIORITIES, index=2, key="cm_prio")
        with cc2:
            new_desc    = st.text_area("Description", height=100, key="cm_desc")
            new_records = st.number_input("Affected Records", min_value=0, value=0, key="cm_recs")
            new_cols    = st.text_input("Affected Columns", placeholder="e.g. email, phone", key="cm_cols")

        if st.button("✅ Create Case", type="primary", key="cm_create"):
            if new_title.strip():
                c = create_case(
                    title=new_title.strip(), case_type=new_type, priority=new_prio,
                    description=new_desc, affected_records=new_records,
                    affected_columns=new_cols, source="Manual",
                )
                st.success(f"✅ Case **{c['case_id']}** created!")
                st.rerun()
            else:
                st.warning("Please enter a case title.")

    if not cases:
        st.info("No cases yet. Create one above or run the Duplicate Studio.")
        return

    st.divider()
    st.markdown("### 📋 All Cases")
    f1, f2, f3 = st.columns(3)
    filt_status = f1.selectbox("Filter by Status",   ["All"] + _CASE_STATUSES,   key="cm_filt_st")
    filt_prio   = f2.selectbox("Filter by Priority", ["All"] + _CASE_PRIORITIES, key="cm_filt_pr")
    filt_type   = f3.selectbox("Filter by Type",     ["All"] + _CASE_TYPES,      key="cm_filt_tp")

    filtered = [c for c in cases
                if (filt_status == "All" or c["status"] == filt_status)
                and (filt_prio == "All" or c["priority"] == filt_prio)
                and (filt_type == "All" or c["type"] == filt_type)]

    if not filtered:
        st.info("No cases match the current filters.")
        return

    for case in sorted(filtered, key=lambda c: c["created_at"], reverse=True):
        label = f"**{case['case_id']}** — {case['title']}  | {case['status']} | {case['priority']}"
        with st.expander(label, expanded=False):
            d1, d2, d3, d4 = st.columns(4)
            d1.markdown(f"**Type:** {case['type']}")
            d2.markdown(f"**Priority:** {case['priority']}")
            d3.markdown(f"**Records:** {case['affected_records']}")
            d4.markdown(f"**Source:** {case['source']}")
            if case["description"]:
                st.markdown(f"**Description:** {case['description']}")
            if case["affected_columns"]:
                st.markdown(f"**Columns:** {case['affected_columns']}")
            st.markdown(f"**Created:** {case['created_at']}  |  **Updated:** {case['updated_at']}")
            if case["resolved_at"]:
                st.markdown(f"**Resolved:** {case['resolved_at']}")
            st.markdown("---")
            u1, u2, u3 = st.columns([1, 1, 1])
            with u1:
                new_st = st.selectbox("Update Status", _CASE_STATUSES,
                                      index=_CASE_STATUSES.index(case["status"]),
                                      key=f"cm_st_{case['case_id']}")
            with u2:
                note = st.text_input("Note", key=f"cm_note_{case['case_id']}")
            with u3:
                st.markdown("<div style='height:1.6rem'></div>", unsafe_allow_html=True)
                if st.button("💾 Update", key=f"cm_upd_{case['case_id']}"):
                    if new_st != case["status"]:
                        update_case_status(case["case_id"], new_st, note)
                        st.success(f"Updated to **{new_st}**")
                        st.rerun()
            if case["history"]:
                st.markdown("**📜 Audit Trail:**")
                for h in reversed(case["history"]):
                    st.markdown(
                        f"<div style='padding:0.3rem 0.8rem;margin:0.2rem 0;"
                        f"background:#f7f4fc;border-left:3px solid #7c3aed;border-radius:4px;"
                        f"font-size:0.85rem;'>"
                        f"<strong>{h['ts']}</strong> — {h['action']} <em>({h['by']})</em>"
                        f"</div>", unsafe_allow_html=True,
                    )


# ══════════════════════════════════════════════════════════════════════════
#  TAB: DYNAMIC DUPLICATE STUDIO  ← NEW
# ══════════════════════════════════════════════════════════════════════════

def _render_dynamic_duplicate_studio():
    st.markdown("### 🔬 Dynamic Duplicate Studio")
    UIComponents.render_action_hint_bar(
        title="How it works",
        message="Load data, profile columns, choose <strong>match mode</strong>, "
                "configure columns & threshold, then <strong>Run Detection</strong>. "
                "Cases are created automatically for every duplicate group found.",
        color="#a78bfa",
    )

    # ── Step 1: Data source ────────────────────────────────────────────────
    st.markdown("#### Step 1 — Load Data")
    source_option = st.radio(
        "Data Source",
        ["Upload new file", "Use DQ results (if available)"],
        horizontal=True, key="studio_src",
    )

    source_df: Optional[pd.DataFrame] = None

    if source_option == "Upload new file":
        dup_file = st.file_uploader(
            "Upload dataset",
            type=AppConfig.SUPPORTED_DATA_FORMATS,
            key="studio_upload",
        )
        if dup_file:
            from modules.data_io_core import FileLoaderService
            loader   = FileLoaderService()
            tmp_path = AppConfig.TEMP_DIR / dup_file.name
            tmp_path.parent.mkdir(parents=True, exist_ok=True)
            tmp_path.write_bytes(dup_file.getbuffer())
            source_df = loader.load_dataframe(tmp_path)
            st.info(f"✅ Loaded **{len(source_df):,}** records · **{len(source_df.columns)}** columns")
    else:
        dq_df = st.session_state.get("dq_results_df")
        if dq_df is not None:
            skip = {"Issues", "Count of issues", "Failed_Rules", "Failed_Columns", "Issue categories"}
            source_df = dq_df[[c for c in dq_df.columns if not c.startswith("_") and c not in skip]].copy()
            st.info(f"✅ Using DQ results: **{len(source_df):,}** records · **{len(source_df.columns)}** columns")
        else:
            st.warning("⚠️ No DQ results available. Upload a file or run a DQ Assessment first.")
            return

    if source_df is None or source_df.empty:
        return

    st.session_state["dup_source_df"] = source_df
    all_cols = [c for c in source_df.columns if not c.startswith("_")]

    # ── Step 2: Column Profiler & Recommendations ──────────────────────────
    st.divider()
    st.markdown("#### Step 2 — Column Profile & Key Recommendations")

    if st.button("🔍 Profile Columns", key="studio_profile_btn"):
        with st.spinner("Profiling columns…"):
            profile_df = profile_columns(source_df)
            st.session_state["studio_profile"] = profile_df
        st.rerun()

    profile_df = st.session_state.get("studio_profile")
    if profile_df is not None:
        st.markdown(
            "<p style='font-size:0.87rem;color:#57534e;'>Column profile — use "
            "<strong>🟢 Strong</strong> identifiers as primary match keys.</p>",
            unsafe_allow_html=True,
        )
        st.dataframe(profile_df, use_container_width=True, hide_index=True)

        strong_cols = profile_df[profile_df["Recommendation"].str.startswith("🟢")]["Column"].tolist()
        if strong_cols:
            st.success(f"**Recommended match keys (Strong identifiers):** {', '.join(strong_cols)}")

    # ── Step 3: Match Mode & Configuration ────────────────────────────────
    st.divider()
    st.markdown("#### Step 3 — Match Mode & Configuration")

    mode = st.radio(
        "Detection Mode",
        ["Exact (Single Column)", "Exact (Multi-Column Combination)", "Fuzzy (Single Column)"],
        horizontal=True, key="studio_mode",
    )
    st.session_state["studio_match_mode"] = mode

    fuzzy      = False
    threshold  = 0.85
    match_cols: List[str] = []

    if mode == "Exact (Single Column)":
        col_choice = st.selectbox("Match Column", all_cols, key="studio_col_single")
        match_cols = [col_choice]
        st.caption("Records with identical values in the selected column will be grouped.")

    elif mode == "Exact (Multi-Column Combination)":
        match_cols = st.multiselect(
            "Match Columns (all must match)",
            options=all_cols,
            default=[],
            key="studio_col_multi",
            help="Records matching on ALL selected columns are grouped as duplicates.",
        )
        if not match_cols:
            st.info("👆 Select at least two columns for combination matching.")
            return
        st.caption(f"Composite key: {' | '.join(match_cols)}")

    else:  # Fuzzy
        col_choice = st.selectbox("Match Column (Fuzzy)", all_cols, key="studio_col_fuzzy")
        match_cols = [col_choice]
        threshold  = st.slider(
            "Similarity Threshold", min_value=0.50, max_value=1.00,
            value=0.85, step=0.01, key="studio_threshold",
            help="Records with similarity ≥ threshold will be grouped. 1.0 = exact.",
        )
        fuzzy = True
        st.caption(
            f"Values with ≥ {threshold:.0%} similarity on **{col_choice}** will be grouped. "
            "Uses SequenceMatcher (character-level ratio)."
        )
        if len(source_df) > 5000:
            st.warning(
                "⚠️ Fuzzy matching on large datasets can be slow. "
                "Consider filtering to < 5,000 rows or sampling first."
            )

    st.session_state["dup_match_columns"] = match_cols

    # ── Survivorship rule (select before running) ─────────────────────────
    surv_strategy = st.selectbox(
        "Survivorship Rule (for auto golden record generation)",
        _SURVIVORSHIP_RULES, key="studio_surv",
    )

    # ── Run ────────────────────────────────────────────────────────────────
    st.divider()
    _, col_btn, _ = st.columns([1, 1, 1])
    with col_btn:
        run_btn = st.button("🚀 Run Duplicate Detection", type="primary",
                            use_container_width=True, key="studio_run")

    if run_btn:
        if not match_cols:
            st.warning("Please select at least one match column.")
            return
        with st.spinner("Detecting duplicates…"):
            dup_df = detect_duplicates(source_df, match_cols, fuzzy=fuzzy, threshold=threshold)
            st.session_state["dup_groups"] = dup_df
            dup_count  = int(dup_df["_is_duplicate"].sum())
            grp_count  = dup_df[dup_df["_is_duplicate"]]["_dup_group_id"].nunique() if dup_count else 0
            match_type = "Fuzzy" if fuzzy else "Exact"

            # Auto-create cases for each group
            n_cases = _auto_create_cases_for_dup_groups(dup_df, match_cols, match_type)

            # Auto golden records
            if dup_count > 0:
                golden_df, discards_df = build_golden_records_df(dup_df, surv_strategy)
                st.session_state["cm_golden_df"]   = golden_df
                st.session_state["cm_discards_df"] = discards_df

        st.rerun()

    # ── Results ────────────────────────────────────────────────────────────
    dup_df = st.session_state.get("dup_groups")
    if dup_df is None or dup_df.empty:
        return

    dup_only  = dup_df[dup_df["_is_duplicate"]]
    dup_count = len(dup_only)
    grp_count = dup_only["_dup_group_id"].nunique() if dup_count > 0 else 0

    st.divider()

    # ── Duplicate Analytics Dashboard ─────────────────────────────────────
    st.markdown("### 📊 Duplicate Analytics Dashboard")

    m1, m2, m3, m4, m5 = st.columns(5)
    m1.metric("Total Records",     f"{len(dup_df):,}")
    m2.metric("Duplicate Records", f"{dup_count:,}")
    m3.metric("Duplicate Groups",  f"{grp_count:,}")
    m4.metric("Unique Records",    f"{len(dup_df) - dup_count:,}")
    exact_cnt = int((dup_only.get("_match_type", pd.Series(dtype=str)) == "Exact").sum()) if "_match_type" in dup_only.columns else 0
    fuzzy_cnt = dup_count - exact_cnt
    m5.metric("Fuzzy Matches",     f"{fuzzy_cnt:,}")

    if dup_count == 0:
        st.success("🎉 No duplicates found! All records are unique on the selected configuration.")
        return

    UIComponents.render_micro_progress(
        int((len(dup_df) - dup_count) / len(dup_df) * 100), "#10b981", "#34d399"
    )

    # Analytics charts
    analytics = _dup_analytics_charts_png(dup_df)
    bar_img   = _dup_group_bar_png(dup_df)
    chart_imgs = {k: v for k, v in analytics.items() if v}
    if bar_img:
        chart_imgs["group_bar"] = bar_img

    if chart_imgs:
        st.divider()
        chart_cols = st.columns(min(len(chart_imgs), 3))
        for i, (key, img) in enumerate(chart_imgs.items()):
            with chart_cols[i % 3]:
                labels = {"match_type": "Duplicate Type Distribution",
                          "fuzzy_dist": "Fuzzy Similarity Distribution",
                          "group_bar":  "Duplicate Groups (Top 20)"}
                st.markdown(f"**{labels.get(key, key)}**")
                st.image(img, use_container_width=True)

    # Most duplicated columns (inferred from match_columns in cases)
    st.divider()
    dup_cases = [c for c in st.session_state["cases"]
                 if c.get("type") == "Duplicate Records" and c.get("source") == "Dynamic Duplicate Studio"]
    if dup_cases:
        col_freq: Dict[str, int] = defaultdict(int)
        for c in dup_cases:
            for col in c.get("affected_columns", "").split(","):
                col = col.strip()
                if col:
                    col_freq[col] += 1
        if col_freq:
            st.markdown("##### Most Duplicated Columns")
            col_df = pd.DataFrame(
                sorted(col_freq.items(), key=lambda x: x[1], reverse=True),
                columns=["Column", "Groups involving this column"],
            )
            st.dataframe(col_df, use_container_width=True, hide_index=True)

    # ── Browse groups ──────────────────────────────────────────────────────
    st.divider()
    st.markdown("### 📋 Browse Duplicate Groups")
    group_ids = sorted(dup_only["_dup_group_id"].unique())
    sel_grp   = st.selectbox("Select Duplicate Group", group_ids, key="studio_sel_grp")

    if sel_grp:
        grp = dup_df[dup_df["_dup_group_id"] == sel_grp]
        display_cols = [c for c in grp.columns if not c.startswith("_")] + [
            "_completeness", "_dup_group_id", "_match_type", "_similarity_score"
        ]
        display_cols = [c for c in display_cols if c in grp.columns]
        st.markdown(f"**Group {sel_grp}** — {len(grp)} records")
        styled = grp[display_cols].style.background_gradient(
            subset=["_completeness"] if "_completeness" in grp.columns else [],
            cmap="Greens", vmin=0, vmax=100,
        )
        st.dataframe(styled, use_container_width=True, hide_index=True)


# ══════════════════════════════════════════════════════════════════════════
#  TAB: GOLDEN RECORDS
# ══════════════════════════════════════════════════════════════════════════

def _render_golden_records_tab():
    st.markdown("### 🏆 Golden Record Identification")
    UIComponents.render_action_hint_bar(
        title="Survivorship logic",
        message="Select a <strong>strategy</strong> to pick the best record from each duplicate group. "
                "The system scores each candidate and highlights the winner.",
        color="#10b981",
    )

    dup_df = st.session_state.get("dup_groups")
    if dup_df is None or dup_df.empty or not dup_df["_is_duplicate"].any():
        st.markdown("""
        <div style="text-align:center;padding:3rem 0;">
            <div style="font-size:3.5rem;margin-bottom:1rem;">🔎</div>
            <h3 style="color:#6d28d9;">No Duplicates Detected Yet</h3>
            <p style="color:#57534e;">
                Use the <strong>Dynamic Duplicate Studio</strong> tab to detect duplicate groups first.
            </p>
        </div>
        """, unsafe_allow_html=True)
        return

    strategy = st.selectbox("Survivorship Strategy", _SURVIVORSHIP_RULES, key="cm_surv_strat")

    _, col_btn, _ = st.columns([1, 1, 1])
    with col_btn:
        run_golden = st.button("🏆 Identify Golden Records", type="primary",
                               use_container_width=True, key="cm_run_golden")

    if run_golden:
        with st.spinner("Identifying golden records…"):
            golden_df, discards_df = build_golden_records_df(dup_df, strategy)
            st.session_state["cm_golden_df"]   = golden_df
            st.session_state["cm_discards_df"] = discards_df
            st.rerun()

    golden_df   = st.session_state.get("cm_golden_df")
    discards_df = st.session_state.get("cm_discards_df")

    if golden_df is None:
        return

    st.divider()
    st.markdown("### 📊 Golden Record Results")
    dup_only  = dup_df[dup_df["_is_duplicate"]]
    n_groups  = dup_only["_dup_group_id"].nunique()
    n_discards= len(discards_df) if discards_df is not None else 0

    m1, m2, m3, m4 = st.columns(4)
    m1.metric("Duplicate Records", f"{len(dup_only):,}")
    m2.metric("Golden Records",    f"{n_groups:,}")
    m3.metric("Records Discarded", f"{n_discards:,}")
    m4.metric("Final Clean Dataset", f"{len(golden_df):,}")
    UIComponents.render_micro_progress(100, "#10b981", "#34d399")
    st.divider()

    v1, v2 = st.columns(2)
    with v1:
        st.markdown("#### Golden vs Discarded")
        pie_img = _golden_vs_discard_pie_png(n_groups, n_discards)
        if pie_img:
            st.image(pie_img, use_container_width=True)
    with v2:
        st.markdown("#### Data Reduction Summary")
        st.markdown(
            f"""
            <div style="background:#f7f4fc;border-radius:12px;padding:1.5rem;border:1px solid #e8e2f5;margin-top:0.5rem;">
                <div style="font-size:1.1rem;color:#3b1f72;font-weight:600;margin-bottom:0.8rem;">
                    Strategy: <span style="color:#7c3aed;">{strategy}</span>
                </div>
                <div style="display:flex;gap:1.5rem;flex-wrap:wrap;">
                    <div>
                        <div style="font-size:0.8rem;color:#6b5f82;">Original</div>
                        <div style="font-size:1.5rem;font-weight:700;color:#1a1028;">{len(dup_df):,}</div>
                    </div>
                    <div style="font-size:1.5rem;color:#7c3aed;align-self:center;">→</div>
                    <div>
                        <div style="font-size:0.8rem;color:#6b5f82;">Clean Dataset</div>
                        <div style="font-size:1.5rem;font-weight:700;color:#10b981;">{len(golden_df):,}</div>
                    </div>
                    <div>
                        <div style="font-size:0.8rem;color:#6b5f82;">Removed</div>
                        <div style="font-size:1.5rem;font-weight:700;color:#ef4444;">{n_discards:,}</div>
                    </div>
                </div>
            </div>
            """, unsafe_allow_html=True,
        )

    st.divider()
    st.markdown("### 🔬 Group-by-Group Comparison")
    group_ids = sorted(dup_only["_dup_group_id"].unique())
    sel_grp   = st.selectbox("Select Duplicate Group", group_ids, key="cm_golden_grp")

    if sel_grp:
        grp = dup_df[dup_df["_dup_group_id"] == sel_grp].copy()
        grp["_is_golden"] = grp.index.isin(set(golden_df.index))
        display_cols = [c for c in grp.columns if not c.startswith("_") or c in ("_completeness", "_is_golden")]
        display_cols = [c for c in display_cols if c in grp.columns]
        st.markdown(f"**Group {sel_grp}** — {len(grp)} records")

        def _highlight_golden(row):
            base = ["background-color: #dcfce7; font-weight: bold"] if row.get("_is_golden") else ["background-color: #fee2e2; opacity: 0.7"]
            return base * len(row)

        styled = grp[display_cols].style.apply(_highlight_golden, axis=1)
        st.dataframe(styled, use_container_width=True, hide_index=True)
        st.markdown(
            '<div style="font-size:0.82rem;color:#57534e;">'
            '🟢 <strong>Green</strong> = Golden Record &nbsp;|&nbsp; 🔴 <strong>Red</strong> = Discarded</div>',
            unsafe_allow_html=True,
        )

    st.divider()
    st.markdown("### ✅ Clean Golden Dataset Preview")
    display_cols = [c for c in golden_df.columns if not c.startswith("_")]
    st.dataframe(golden_df[display_cols].head(100), use_container_width=True, hide_index=True)
    st.info(f"Showing first 100 of {len(golden_df):,} records")


# ══════════════════════════════════════════════════════════════════════════
#  TAB: REPORTS & EXPORT
# ══════════════════════════════════════════════════════════════════════════

def _render_reports_tab():
    st.markdown("### 📥 Reports & Export")
    UIComponents.render_action_hint_bar(
        title="Download",
        message="Export <strong>Case Summary</strong>, <strong>Duplicate Groups</strong>, "
                "<strong>Golden Records</strong>, and <strong>Discards</strong> as a multi-sheet Excel workbook.",
        color="#60a5fa",
    )

    cases       = st.session_state["cases"]
    dup_df      = st.session_state.get("dup_groups")
    golden_df   = st.session_state.get("cm_golden_df")
    discards_df = st.session_state.get("cm_discards_df")

    if not cases and dup_df is None:
        st.info("No data to export yet. Create cases or run the Duplicate Studio first.")
        return

    st.divider()
    _, col_btn, _ = st.columns([1, 1, 1])
    with col_btn:
        gen = st.button("📊 Generate Excel Report", type="primary", use_container_width=True, key="cm_gen_xl")

    if gen:
        with st.spinner("Building Excel report…"):
            xl_bytes = build_case_excel(cases, dup_df, golden_df, discards_df)
            st.session_state["cm_excel_bytes"] = xl_bytes
            st.rerun()

    xl_bytes = st.session_state.get("cm_excel_bytes")
    if xl_bytes:
        ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        st.download_button(
            "📥 Download Case Management Report",
            data=xl_bytes,
            file_name=f"Case_Management_Report_{ts}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True, key="cm_dl_xl",
        )
        UIComponents.render_hint_chip(
            "Multi-sheet workbook",
            tip="Includes: Case Summary, Duplicate Groups, Golden Records, Discards",
            icon="📊",
        )

    st.divider()

    if golden_df is not None and not golden_df.empty:
        st.markdown("#### 🏆 Download Clean Golden Dataset")
        display_cols = [c for c in golden_df.columns if not c.startswith("_")]
        csv_buf = BytesIO()
        golden_df[display_cols].to_csv(csv_buf, index=False)
        csv_buf.seek(0)
        st.download_button(
            "📄 Download Golden Records (CSV)",
            data=csv_buf.getvalue(),
            file_name=f"Golden_Records_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
            mime="text/csv", use_container_width=True, key="cm_dl_golden_csv",
        )

    dq_xl_path = st.session_state.get("dq_excel_path")
    if dq_xl_path and dq_xl_path.exists():
        st.divider()
        st.markdown("#### 📊 Linked DQ Assessment Report")
        with open(dq_xl_path, "rb") as f:
            st.download_button(
                "📊 Download DQ Assessment Report",
                data=f.read(),
                file_name=dq_xl_path.name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True, key="cm_dl_dq",
            )
