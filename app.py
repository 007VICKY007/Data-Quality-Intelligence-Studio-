# ── stdlib ─────────────────────────────────────────────────────────────────
import traceback
import datetime
from io import BytesIO
from pathlib import Path

# ── third-party ────────────────────────────────────────────────────────────
import streamlit as st
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from matplotlib.patches import Wedge


# ══════════════════════════════════════════════════════════════════════════
#  DQ ENGINE MODULES  (merged — replaces old split imports)
# ══════════════════════════════════════════════════════════════════════════
from modules.config          import AppConfig
from modules.case_management import page_case_management, init_case_management_state
from modules.ui_components   import UIComponents


# ══════════════════════════════════════════════════════════════════════════
#  PATCH — Fix broken HTML in UIComponents methods
#  The original render_lottie_upload produces raw ' style="color" text
#  visible in the UI. These replacements use safe, working HTML.
# ══════════════════════════════════════════════════════════════════════════
def _render_lottie_upload_fixed(caption: str = "Upload both files above to begin") -> None:
    st.markdown(
        """
        <div class="lottie-slot">
            <div class="lottie-frame lottie-upload-fallback"></div>
            <p class="lottie-caption">{caption}</p>
        </div>
        """.replace("{caption}", caption),
        unsafe_allow_html=True,
    )

def _render_arrow_down_fixed() -> None:
    st.markdown(
        '<div class="guidance-arrow-down">⬇</div>',
        unsafe_allow_html=True,
    )

def _render_upload_hint_fixed(kind: str = "dataset") -> None:
    if kind == "dataset":
        label = "📂 Master Dataset"
        tip   = "CSV, Excel (.xlsx/.xls/.xlsm), JSON, Parquet, ODS or XML"
    else:
        label = "📜 Rules / Rulebook"
        tip   = "CSV/Excel with column_name, rule, dimension, message — or a JSON rulebook"
    st.markdown(
        '<p style="font-size:0.82rem;color:#64748b;margin-bottom:0.3rem;">'
        + label + " &nbsp;·&nbsp; " + tip
        + "</p>",
        unsafe_allow_html=True,
    )

def _render_results_header_fixed(score: float) -> None:
    if score >= 80:
        cls, emoji, label = "dq-score-excellent", "🏆", "Excellent"
    elif score >= 60:
        cls, emoji, label = "dq-score-good",      "✅", "Good"
    elif score >= 40:
        cls, emoji, label = "dq-score-fair",      "⚠️", "Fair"
    else:
        cls, emoji, label = "dq-score-poor",      "❌", "Poor"
    st.markdown(
        '<div class="' + cls + '">'
        + '<h2 style="margin:0;">' + emoji + ' ' + label + ' — ' + f'{score:.1f}%' + '</h2>'
        + '</div>',
        unsafe_allow_html=True,
    )

# Patch UIComponents with fixed methods
UIComponents.render_lottie_upload  = _render_lottie_upload_fixed
UIComponents.render_arrow_down     = _render_arrow_down_fixed
UIComponents.render_upload_hint    = _render_upload_hint_fixed
UIComponents.render_results_header = _render_results_header_fixed

# data_quality_core  ←  dq_engine + rule_executor + rulebook_builder
from modules.data_quality_core import (
    RulebookBuilderService,
    RuleExecutorEngine,
    DataQualityEngine,
)

# reporting_core     ←  scoring_engine + report_generator
from modules.reporting_core import (
    ScoringService,
    ExcelReportGenerator,
)

# data_io_core       ←  file_loader + utils
from modules.data_io_core import (
    FileLoaderService,
    setup_directories,
    save_uploaded_file,
    clean_temp_directory,
)


# ══════════════════════════════════════════════════════════════════════════
#  DATA MATURITY MODULES
# ══════════════════════════════════════════════════════════════════════════
from DataMaturity.config import (
    UNIQU_PURPLE, UNIQU_MAGENTA, UNIQU_LAVENDER,
    UNIQU_LIGHT_BG, UNIQU_TEXT, UNIQU_GREY,
    RATING_LABELS, RATING_TO_SCORE,
    DEFAULT_MASTER_OBJECTS, MATURITY_DIMS, QUESTION_BANK,
)

from DataMaturity.helpers import (
    dq_score_to_maturity_level,
    init_maturity_state,
    build_question_df,
    sync_response_tables,
    autofill_dq_dimension,
    compute_all_scores,
    validate_responses,
    to_excel_bytes,
)

from DataMaturity.visualizations   import render_slide_png
from DataMaturity.report_generator import build_pdf_bytes

# ══════════════════════════════════════════════════════════════════════════
#  EXTERNAL CSS — assets/styles.css
# ══════════════════════════════════════════════════════════════════════════
def load_css():
    """Load external stylesheet from assets folder."""
    try:
        with open("assets/styles.css", encoding="utf-8") as f:
            st.markdown(f"<style>{f.read()}</style>", unsafe_allow_html=True)
    except FileNotFoundError:
        st.warning("⚠️ styles.css not found in assets/ folder — place it at assets/styles.css")


# ══════════════════════════════════════════════════════════════════════════
#  FORCE FIX — DATA EDITOR DROPDOWN DARK THEME
# ══════════════════════════════════════════════════════════════════════════
st.markdown("""
<style>
div[data-baseweb="popover"],
div[data-baseweb="popover"] > div,
div[data-baseweb="menu"] {
    background: #0f172a !important;
    border: 1px solid #334155 !important;
}
div[data-baseweb="popover"] ul,
div[data-baseweb="menu"] ul {
    background: #0f172a !important;
}
div[data-baseweb="popover"] li *,
div[data-baseweb="menu"] li *,
div[data-baseweb="popover"] [role="option"] *,
div[data-baseweb="menu"] [role="option"] * {
    color: #f8fafc !important;
    -webkit-text-fill-color: #f8fafc !important;
}
div[data-baseweb="popover"] [role="option"],
div[data-baseweb="menu"] [role="option"] {
    background: #1e293b !important;
    color: #f8fafc !important;
    font-weight: 600 !important;
}
div[data-baseweb="popover"] [role="option"]:hover {
    background: #334155 !important;
}
div[data-baseweb="popover"] [aria-selected="true"] {
    background: rgba(124,58,237,0.35) !important;
    color: #e9d5ff !important;
}
div[data-baseweb="popover"] [data-highlighted] {
    background: rgba(96,165,250,0.25) !important;
}
</style>
""", unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════
#  GDG LIGHT THEME
#  Streamlit's Glide Data Grid reads --gdg-* vars from :root only,
#  so we inject them via st.markdown on every page that uses data_editor
# ══════════════════════════════════════════════════════════════════════════
_GDG_LIGHT_STYLE = """
<style>
/* ── Glide Data Grid: peaceful white/lavender theme ── */
:root,
[data-testid="stDataEditor"],
[data-testid="stDataEditor"] > div {
    --gdg-bg-cell:                 #ffffff !important;
    --gdg-bg-cell-medium:          #f7f4fc !important;
    --gdg-bg-header:               #ede8f7 !important;
    --gdg-bg-header-has-focus:     #e0d9f2 !important;
    --gdg-bg-header-hovered:       #d4cced !important;
    --gdg-border-color:            #e8e2f5 !important;
    --gdg-horizontal-border-color: #e8e2f5 !important;
    --gdg-accent-color:            #7c3aed !important;
    --gdg-accent-light:            rgba(124,58,237,0.10) !important;
    --gdg-text-dark:               #1a1028 !important;
    --gdg-text-medium:             #3b2f54 !important;
    --gdg-text-light:              #6b5f82 !important;
    --gdg-text-header:             #3b1f72 !important;
    --gdg-text-header-selected:    #1a0a40 !important;
    --gdg-cell-text-color:         #1a1028 !important;
    --gdg-header-font-style:       700 13px -apple-system, sans-serif !important;
    --gdg-base-font-style:         500 13px -apple-system, sans-serif !important;
}
[data-testid="stDataEditor"] canvas {
    background-color: #ffffff !important;
}
[data-testid="stDataEditor"] .dvn-scroller,
[data-testid="stDataEditor"] .dvn-scroll-inner,
[data-testid="stDataEditor"] > div,
[data-testid="stDataEditor"] > div > div {
    background: #ffffff !important;
}
[data-testid="stDataEditor"] ::-webkit-scrollbar        { width:7px; height:7px; }
[data-testid="stDataEditor"] ::-webkit-scrollbar-track  { background:#f3f0fa !important; }
[data-testid="stDataEditor"] ::-webkit-scrollbar-thumb  {
    background: #c4b5fd !important;
    border-radius: 4px;
}
[data-testid="stDataEditor"] ::-webkit-scrollbar-thumb:hover { background:#a78bfa !important; }
[data-testid="stDataEditor"] ::-webkit-scrollbar-corner      { background:#f3f0fa !important; }
[data-testid="stDataEditor"] input {
    background: #ffffff !important;
    color: #1a1028 !important;
    -webkit-text-fill-color: #1a1028 !important;
    border: 2px solid #7c3aed !important;
    border-radius: 5px !important;
    box-shadow: 0 0 0 3px rgba(124,58,237,0.12) !important;
}
.gdg-overlay-editor {
    background: #ffffff !important;
    border: 1.5px solid #d8d0ed !important;
    border-radius: 8px !important;
    box-shadow: 0 8px 28px rgba(91,45,144,0.15) !important;
    color: #1a1028 !important;
}
.gdg-overlay-editor select,
.gdg-overlay-editor select option {
    background: #ffffff !important;
    color: #1a1028 !important;
    -webkit-text-fill-color: #1a1028 !important;
}
</style>
"""


def inject_gdg_light():
    """Call once per page that renders st.data_editor."""
    st.markdown(_GDG_LIGHT_STYLE, unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════
#  SESSION STATE INITIALIZATION
# ══════════════════════════════════════════════════════════════════════════
def _init_state() -> None:
    # Navigation
    if "page" not in st.session_state:
        st.session_state["page"] = "home"

    # DQ results (populated after DQ run; consumed by Maturity)
    for key, default in {
        "dq_score":       None,
        "dq_dim_scores":  None,
        "dq_results_df":  None,
        "dq_object_name": "Customer",
        "dq_excel_path":  None,
    }.items():
        if key not in st.session_state:
            st.session_state[key] = default

    # Maturity state (uses keys defined in DataMaturity/helpers.py)
    init_maturity_state()

    # Policy Hub state
    if "policies" not in st.session_state:
        st.session_state["policies"] = []

    # Case Management state
    if "cases" not in st.session_state:
        st.session_state["cases"] = []


# ══════════════════════════════════════════════════════════════════════════
#  UTILITY FUNCTIONS
# ══════════════════════════════════════════════════════════════════════════
def get_timestamp_filename(prefix: str, extension: str) -> str:
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"{prefix}_{timestamp}.{extension}"


# ══════════════════════════════════════════════════════════════════════════
#  VISUALIZATION FUNCTIONS
# ══════════════════════════════════════════════════════════════════════════
def _gauge_png(score: float) -> bytes:
    fig, ax = plt.subplots(figsize=(5, 3.2), dpi=150)
    fig.patch.set_facecolor('#fafafa')
    ax.set_xlim(0, 1); ax.set_ylim(0, 0.65); ax.axis("off")
    ax.add_patch(Wedge((0.5, 0.05), 0.40, 0, 180, width=0.12,
                       facecolor="#e5e7eb", edgecolor="white", lw=3))
    ang = score / 100 * 180
    col = "#10b981" if score >= 80 else ("#f59e0b" if score >= 60 else "#ef4444")
    ax.add_patch(Wedge((0.5, 0.05), 0.40, 0, ang, width=0.12,
                       facecolor=col, edgecolor="white", lw=3))
    ax.text(0.5, 0.32, f"{score:.1f}%", ha="center", va="center",
            fontsize=28, fontweight="bold", color="#6d28d9", family="sans-serif")
    ax.text(0.5, 0.18, "Overall DQ Score", ha="center", va="center",
            fontsize=11, color="#57534e", family="sans-serif", weight=600)
    buf = BytesIO()
    fig.savefig(buf, format="png", bbox_inches="tight", pad_inches=0.15, facecolor='#fafafa')
    plt.close(fig)
    return buf.getvalue()


def _dim_bar_png(dim_scores: dict) -> bytes | None:
    if not dim_scores:
        return None
    dims   = list(dim_scores.keys())
    scores = [dim_scores[d] for d in dims]
    cols   = ["#10b981" if s >= 80 else ("#f59e0b" if s >= 60 else "#ef4444") for s in scores]
    fig, ax = plt.subplots(figsize=(8, max(3, len(dims) * 0.8)), dpi=140)
    fig.patch.set_facecolor('#fafafa'); ax.set_facecolor('#ffffff')
    bars = ax.barh(dims, scores, color=cols, height=0.6, edgecolor="white", linewidth=2)
    ax.set_xlim(0, 112)
    ax.set_xlabel("DQ Score (%)", color="#1c1917", fontsize=11, weight=600, family="sans-serif")
    ax.tick_params(colors="#44403c", labelsize=10)
    ax.spines[["top", "right", "bottom"]].set_visible(False)
    ax.spines["left"].set_color("#d6d3d1"); ax.spines["left"].set_linewidth(1.5)
    ax.axvline(80, color="#6d28d9", lw=1.5, ls="--", alpha=0.6, label="Excellent (80%)")
    ax.axvline(60, color="#7c3aed", lw=1.5, ls=":",  alpha=0.6, label="Good (60%)")
    ax.legend(fontsize=9, loc="lower right", frameon=True, fancybox=True, shadow=True)
    for bar, sc in zip(bars, scores):
        ax.text(bar.get_width() + 2, bar.get_y() + bar.get_height() / 2,
                f"{sc:.1f}%", va="center", fontsize=10,
                fontweight="bold", color="#1c1917", family="sans-serif")
    plt.tight_layout()
    buf = BytesIO()
    fig.savefig(buf, format="png", bbox_inches="tight", facecolor='#fafafa')
    plt.close(fig)
    return buf.getvalue()


def _mat_bar_png(dim_vals: dict) -> bytes | None:
    if not dim_vals:
        return None
    dims   = list(dim_vals.keys())
    scores = [dim_vals[d] for d in dims]
    cols   = ["#0369a1" if s >= 4 else ("#0ea5e9" if s >= 3 else "#7dd3fc") for s in scores]
    fig, ax = plt.subplots(figsize=(10, max(3, len(dims) * 0.9)), dpi=140)
    fig.patch.set_facecolor('#f0f9ff'); ax.set_facecolor('#ffffff')
    bars = ax.barh(dims, scores, color=cols, height=0.6, edgecolor="white", linewidth=2)
    ax.set_xlim(0, 6.0)
    ax.set_xlabel("Maturity Score (1 = Adhoc  →  5 = Optimised)",
                  color="#0c4a6e", fontsize=11, weight=600, family="sans-serif")
    ax.axvline(3.0, color="#38bdf8", lw=1.5, ls="--", alpha=0.7, label="Defined (3)")
    ax.axvline(4.0, color="#0284c7", lw=1.5, ls="--", alpha=0.7, label="Managed (4)")
    ax.legend(fontsize=9, loc="lower right", frameon=True, fancybox=True, shadow=True)
    ax.tick_params(colors="#0c4a6e", labelsize=10)
    ax.spines[["top", "right", "bottom"]].set_visible(False)
    ax.spines["left"].set_color("#bae6fd"); ax.spines["left"].set_linewidth(2)
    for bar, sc in zip(bars, scores):
        ax.text(bar.get_width() + 0.1, bar.get_y() + bar.get_height() / 2,
                f"{sc:.2f}", va="center", fontsize=11,
                fontweight="bold", color="#0c4a6e", family="sans-serif")
    plt.tight_layout()
    buf = BytesIO()
    fig.savefig(buf, format="png", bbox_inches="tight", facecolor='#f0f9ff')
    plt.close(fig)
    return buf.getvalue()


# ══════════════════════════════════════════════════════════════════════════
#  COMBINED EXCEL (DQ + Maturity)
# ══════════════════════════════════════════════════════════════════════════
def _combined_excel(dq_score: float, dq_dim_scores: dict | None, mat_excel: bytes) -> bytes:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = load_workbook(BytesIO(mat_excel))
    header_fill = PatternFill(start_color="6d28d9", end_color="7c3aed", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)

    ws_dq = wb.create_sheet("DQ Score Summary", 0)
    ws_dq.append(["Metric", "Value"])
    for cell in ws_dq[1]:
        cell.fill = header_fill; cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws_dq.append(["Overall DQ Score (%)",  f"{dq_score:.1f}%"])
    ws_dq.append(["Mapped Maturity Level", dq_score_to_maturity_level(dq_score)])
    if dq_dim_scores:
        for dim, sc in dq_dim_scores.items():
            ws_dq.append([f"DQ – {dim}", f"{sc:.1f}%"])
    ws_dq.column_dimensions['A'].width = 30
    ws_dq.column_dimensions['B'].width = 20

    dq_df = st.session_state.get("dq_results_df")
    if dq_df is not None:
        display_cols = [c for c in dq_df.columns if not c.startswith("_")]
        ws_res = wb.create_sheet("DQ Results", 1)
        ws_res.append(display_cols)
        for cell in ws_res[1]:
            cell.fill = header_fill; cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for _, row in dq_df[display_cols].head(1000).iterrows():
            ws_res.append([str(v) if v is not None else "" for v in row.tolist()])

    out = BytesIO(); wb.save(out); out.seek(0)
    return out.getvalue()

# ══════════════════════════════════════════════════════════════════════════
#  PAGE: HOME
# ══════════════════════════════════════════════════════════════════════════
def page_home():
    import time

    # Background
    st.markdown('<div class="animated-bg"></div>', unsafe_allow_html=True)

    # ─────────────────────────────────────────────────────
    # Typing Header — Tool Name Only
    # ─────────────────────────────────────────────────────
    if "header_typed" not in st.session_state:
        tool_name = "Data Quality Intelligence Studio"
        placeholder = st.empty()
        typed = ""

        for char in tool_name:
            typed += char
            placeholder.markdown(
                f"""
                <h1 style="
                    text-align:center;
                    font-size:2.6rem;
                    font-weight:800;
                    color:#e2e8f0;
                    margin-bottom:0.3rem;
                ">
                    {typed}
                </h1>
                """,
                unsafe_allow_html=True
            )
            time.sleep(0.02)

        st.session_state["header_typed"] = True
    else:
        st.markdown(
            """
            <h1 style="
                text-align:center;
                font-size:2.6rem;
                font-weight:800;
                color:#e2e8f0;
                margin-bottom:0.3rem;
            ">
                Data Quality Intelligence Studio
            </h1>
            """,
            unsafe_allow_html=True
        )

    # ─────────────────────────────────────────────────────
    # Powered By
    # ─────────────────────────────────────────────────────
    st.markdown(
        """
        <p style="
            text-align:center;
            font-size:1rem;
            color:#94a3b8;
            margin-bottom:0.4rem;
        ">
            Powered by
            <span style="color:#60a5fa;font-weight:700;">
                Uniqus Consultech
            </span>
        </p>
        """,
        unsafe_allow_html=True
    )

    # ─────────────────────────────────────────────────────
    # Tagline
    # ─────────────────────────────────────────────────────
    st.markdown(
        """
        <p style="
            text-align:center;
            font-size:1rem;
            color:#cbd5e1;
            max-width:780px;
            margin:auto;
        ">
            Profile, validate, and monitor enterprise data using automated rules,
            AI-driven insights, and dimension-based scoring.
        </p>
        """,
        unsafe_allow_html=True
    )

    st.divider()

    # ─────────────────────────────────────────────────────
    # DQ Completion Banner
    # ─────────────────────────────────────────────────────
    if st.session_state.dq_score is not None:
        sc  = st.session_state.dq_score
        lvl = dq_score_to_maturity_level(sc)

        st.markdown('<div class="banner success">', unsafe_allow_html=True)
        col1, col2 = st.columns([3, 1])

        with col1:
            st.markdown(
                f"""
                ✅ **DQ Assessment Completed**  
                **Score:** {sc:.1f}% | **Level:** {lvl} |  
                **Object:** {st.session_state.dq_object_name}
                """
            )

        with col2:
            if st.button("View Results →", use_container_width=True):
                st.session_state["page"] = "dq"
                st.rerun()

        st.markdown('</div>', unsafe_allow_html=True)
        st.divider()

    # ─────────────────────────────────────────────────────
    # Solutions Workspace — FULL WIDTH CARDS
    # ─────────────────────────────────────────────────────
    st.markdown(
        '<h2 style="text-align:center;margin-bottom:2rem;">Solutions Workspace</h2>',
        unsafe_allow_html=True,
    )

    col1, col2 = st.columns(2, gap="large")

    with col1:
        st.markdown(
            """
            <div class="feature-card large">
                <div class="feature-card-icon">🔍</div>
                <h3>Data Quality Assessment</h3>
                <p>
                Upload dataset and rules to generate automated DQ scores,
                column analysis, dimension scoring and enterprise reports.
                </p>
            </div>
            """,
            unsafe_allow_html=True,
        )

        if st.button("Start DQ Assessment →", use_container_width=True):
            st.session_state["page"] = "dq"
            st.rerun()

    with col2:
        st.markdown(
            """
            <div class="feature-card large">
                <div class="feature-card-icon">📈</div>
                <h3>Data Maturity Assessment</h3>
                <p>
                Evaluate DAMA maturity dimensions, generate executive visuals,
                PDF reports and Excel outputs.
                </p>
            </div>
            """,
            unsafe_allow_html=True,
        )

        if st.button("Start Maturity Assessment →", use_container_width=True):
            st.session_state["page"] = "maturity"
            st.rerun()

    st.markdown("<br>", unsafe_allow_html=True)

    col3, col4 = st.columns(2, gap="large")

    with col3:
        st.markdown(
            """
            <div class="feature-card large">
                <div class="feature-card-icon">📋</div>
                <h3>Policy Hub</h3>
                <p>
                Central governance repository for policy workflows,
                approval tracking and compliance monitoring.
                </p>
            </div>
            """,
            unsafe_allow_html=True,
        )

        if st.button("Open Policy Hub →", use_container_width=True):
            st.session_state["page"] = "policy"
            st.rerun()

    with col4:
        st.markdown(
            """
            <div class="feature-card large">
                <div class="feature-card-icon">🎯</div>
                <h3>Case Management</h3>
                <p>
                Track and resolve data quality issues with ownership,
                SLA tracking and audit trails.
                </p>
            </div>
            """,
            unsafe_allow_html=True,
        )

        if st.button("Open Case Management →", use_container_width=True):
            st.session_state["page"] = "cases"
            st.rerun()

    st.divider()
# ══════════════════════════════════════════════════════════════════════════
#  PAGE: DQ ASSESSMENT
# ══════════════════════════════════════════════════════════════════════════
def page_dq():
    with st.sidebar:
        st.markdown("### 🧭 Navigation")
        if st.button("🏠 Home",      use_container_width=True, key="dq_home"):
            st.session_state["page"] = "home"; st.rerun()
        if st.button("📈 Maturity",  use_container_width=True, key="dq_maturity"):
            st.session_state["page"] = "maturity"; st.rerun()
        if st.button("📋 Policies",  use_container_width=True, key="dq_policy"):
            st.session_state["page"] = "policy"; st.rerun()
        if st.button("🎯 Cases",     use_container_width=True, key="dq_cases"):
            st.session_state["page"] = "cases"; st.rerun()
        st.divider()
        UIComponents.render_sidebar()

    # ── Page header ────────────────────────────────────────────────────────
    st.markdown("# 🔍 Data Quality Assessment")
    st.markdown(
        "Upload your master dataset and rules configuration to generate "
        "comprehensive DQ reports with detailed scoring and analysis."
    )
    st.divider()

    # ── File Upload section ────────────────────────────────────────────────
    st.markdown("### 📁 Input Files")
    col1, col2 = st.columns(2)

    with col1:
        UIComponents.render_upload_hint("dataset")
        data_file = st.file_uploader(
            "New Master Dataset (CSV, Excel, or JSON)",
            type=AppConfig.SUPPORTED_DATA_FORMATS,
            help="CSV, Excel, or JSON format",
            key="dq_data_uploader",
        )

    with col2:
        UIComponents.render_upload_hint("rules")
        rules_file = st.file_uploader(
            "Business Rules Configuration",
            type=AppConfig.SUPPORTED_RULES_FORMATS + ["json"],
            help="CSV, Excel rules sheet, or JSON rulebook",
            key="dq_rules_uploader",
        )

    # ── Empty state ────────────────────────────────────────────────────────
    if not data_file or not rules_file:
        st.markdown('<div style="height:1rem;"></div>', unsafe_allow_html=True)

        _, col_c, _ = st.columns([1, 1.2, 1])
        with col_c:
            UIComponents.render_lottie_upload("Upload both files above to begin")

        UIComponents.render_arrow_down()

        st.markdown("""
        <div class="welcome-steps">
            <div class="welcome-step-card current-step">
                <div class="wsc-number step-1">1</div>
                <span class="wsc-icon animate-upload">📤</span>
                <div class="wsc-title">Upload Your Files</div>
                <p class="wsc-desc">Drop your master dataset (CSV / Excel / JSON)
                   and business rules configuration.</p>
            </div>
            <div class="welcome-step-card">
                <div class="wsc-number step-2">2</div>
                <span class="wsc-icon animate-spin">⚙️</span>
                <div class="wsc-title">Generate Rulebook</div>
                <p class="wsc-desc">Rules are automatically mapped and
                   validation logic is built from your configuration.</p>
            </div>
            <div class="welcome-step-card">
                <div class="wsc-number step-3">3</div>
                <span class="wsc-icon animate-float">📊</span>
                <div class="wsc-title">Get DQ Results</div>
                <p class="wsc-desc">Interactive dashboard with column scores,
                   dimension breakdowns and Excel reports.</p>
            </div>
        </div>
        """, unsafe_allow_html=True)

        st.markdown('<div style="height:1rem;"></div>', unsafe_allow_html=True)

        g1, g2, g3 = st.columns(3)
        with g1:
            UIComponents.render_guidance_card(
                "📁", "Accepted Data Formats",
                "CSV, Excel (.xlsx / .xls / .xlsm), JSON arrays, Parquet, ODS and XML.",
                step_number=1, delay_ms=50,
            )
        with g2:
            UIComponents.render_guidance_card(
                "📜", "Rules File Format",
                "CSV/Excel with columns: column_name, rule, dimension, message — "
                "or upload a pre-built JSON rulebook.",
                step_number=2, delay_ms=150,
            )
        with g3:
            UIComponents.render_guidance_card(
                "📊", "What You Get",
                "Overall DQ score, per-column breakdowns, dimension heatmap "
                "and a multi-sheet Excel report with annexures.",
                step_number=3, delay_ms=250,
            )
        UIComponents.render_file_format_help()
        return

    # ── Both files uploaded ────────────────────────────────────────────────
    UIComponents.render_workflow_tracker(active_step=1)
    UIComponents.render_action_hint_bar(
        title="Files loaded",
        message="Configure the object name below then click "
                "<strong>🚀 Run DQ Assessment</strong> to begin scoring.",
    )
    cfg1, cfg2 = st.columns(2)
    with cfg1:
        obj_name = st.text_input(
            "📌 Master Object / Dataset Name",
            value=st.session_state.get("dq_object_name", "Customer"),
            placeholder="e.g. Customer, Vendor, Material…",
            help="Used to label reports and link results to the Maturity Assessment.",
            key="dq_obj_name_input",
        )
    with cfg2:
        sheet_name = None
        if data_file.name.lower().endswith((".xlsx", ".xls", ".xlsm")):
            loader = FileLoaderService()
            tmp    = AppConfig.TEMP_DIR / data_file.name
            tmp.write_bytes(data_file.getbuffer())
            sheets = loader.get_sheet_names(tmp)
            if len(sheets) > 1:
                sheet_name = st.selectbox("Select Sheet", sheets, key="dq_sheet")

    UIComponents.render_file_format_help()
    st.divider()

    # ── Run Button ─────────────────────────────────────────────────────────
    col1, col2, col3 = st.columns([1, 1, 1])
    with col2:
        run_button = st.button(
            "🚀 Run DQ Assessment",
            type="primary",
            use_container_width=True,
            key="dq_run",
        )

    if not run_button:
        return

    try:
        clean_temp_directory()
        pb   = st.progress(0, text="📂 Saving files...")
        stat = st.empty()

        stat.text("📂 Saving files…"); pb.progress(5, text="📂 Saving files...")
        data_path  = save_uploaded_file(data_file,  AppConfig.TEMP_DIR)
        rules_path = save_uploaded_file(rules_file, AppConfig.TEMP_DIR)

        stat.text("📊 Loading dataset…"); pb.progress(15, text="📊 Loading dataset...")
        loader = FileLoaderService()
        df     = loader.load_dataframe(data_path, sheet_name=sheet_name)
        cols   = list(df.columns)
        st.info(f"✅ Loaded **{len(df):,}** records · **{len(cols)}** columns")

        UIComponents.render_workflow_tracker(active_step=2)
        stat.text("🔧 Building rulebook…"); pb.progress(30, text="🔧 Building rulebook...")
        rb_svc = RulebookBuilderService()
        if rules_file.name.lower().endswith(".json"):
            rulebook = rb_svc.load_json_rulebook(rules_path)
        else:
            rulebook = rb_svc.build_from_rules_dataset(
                loader.load_dataframe(rules_path), cols)

        UIComponents.render_workflow_tracker(active_step=2)
        stat.text("✅ Executing rules…"); pb.progress(50, text="✅ Executing rules...")
        executor = RuleExecutorEngine(df, rulebook)
        results  = executor.execute_all_rules()
        combos   = executor.get_combination_duplicates()

        UIComponents.render_workflow_tracker(active_step=3)
        stat.text("📊 Scoring…"); pb.progress(70, text="📊 Calculating scores...")
        # ScoringService now uses static methods (merged module)
        overall    = ScoringService.calculate_overall_score(results)
        col_scores = ScoringService.calculate_column_scores(results, cols)
        dim_scores = ScoringService.calculate_dimension_scores(results)

        UIComponents.render_workflow_tracker(active_step=4)
        stat.text("💾 Generating Excel report…"); pb.progress(85, text="💾 Generating Excel report...")
        excel_filename = get_timestamp_filename(f"DQ_Report_{obj_name or 'Dataset'}", "xlsx")
        xl_path        = AppConfig.OUTPUT_DIR / excel_filename
        rgen           = ExcelReportGenerator(
            results_df=results, rulebook=rulebook, all_columns=cols,
            column_scores=col_scores, overall_score=overall,
            dimension_scores=dim_scores, duplicate_combinations=combos,
        )
        rgen.generate_report(AppConfig.OUTPUT_DIR)
        rb_path = rgen.save_rulebook_json(AppConfig.OUTPUT_DIR, rulebook)

        pb.progress(100, text="✅ Complete!")
        stat.success("✅ Assessment completed successfully!")

        # Save to session state
        st.session_state["dq_score"]       = overall
        st.session_state["dq_dim_scores"]  = dim_scores
        st.session_state["dq_results_df"]  = results
        st.session_state["dq_object_name"] = obj_name or "Customer"
        st.session_state["dq_excel_path"]  = xl_path

        st.session_state["mat_objects"] = [obj_name] if obj_name else DEFAULT_MASTER_OBJECTS[:]
        autofill_dq_dimension(overall)

        UIComponents.render_results_header(overall)
        st.divider()

        st.markdown("## 📊 Results Dashboard")
        g1, g2 = st.columns([1, 2])
        with g1:
            st.image(_gauge_png(overall), use_container_width=True)
        with g2:
            bar = _dim_bar_png(dim_scores)
            if bar:
                st.image(bar, use_container_width=True)

        UIComponents.render_micro_progress(int(overall), "#10b981" if overall >= 80 else "#f59e0b")
        st.divider()

        UIComponents.render_results_dashboard(overall, results, col_scores, dim_scores)
        st.divider()

        st.markdown("### 📥 Download Reports")
        UIComponents.render_hint_chip(
            "Reports ready", tip="Click to download to your local machine", icon="📥"
        )
        st.markdown('<div style="height:0.5rem;"></div>', unsafe_allow_html=True)

        d1, d2, d3 = st.columns(3)
        with d1:
            default_report_path = AppConfig.OUTPUT_DIR / "DQ_Assessment_Report.xlsx"
            if default_report_path.exists() and not xl_path.exists():
                default_report_path.rename(xl_path)
            if xl_path.exists():
                with open(xl_path, "rb") as f:
                    st.download_button(
                        "📊 DQ Excel Report", data=f.read(),
                        file_name=excel_filename,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                    )
            else:
                st.error("❌ Excel report not found")
        with d2:
            if rb_path and Path(rb_path).exists():
                with open(rb_path, "rb") as f:
                    rb_filename = get_timestamp_filename("Rulebook", "json")
                    st.download_button(
                        "📋 Rulebook JSON", data=f.read(),
                        file_name=rb_filename, mime="application/json",
                        use_container_width=True,
                    )
        with d3:
            st.info(f"✅ {len(cols)} columns analyzed")

        st.divider()
        UIComponents.render_detailed_views(rulebook, results, col_scores, dim_scores)

        st.divider()
        lvl = dq_score_to_maturity_level(overall)
        st.markdown(
            f'<div class="banner">'
            f'💡 DQ Score <span class="purple-text">{overall:.1f}%</span> maps to maturity level '
            f'<span class="purple-text">{lvl}</span>. '
            f'This has been auto-filled in the Data Quality assessment dimension.'
            f'</div>',
            unsafe_allow_html=True,
        )

        UIComponents.render_action_hint_bar(
            title="Next step",
            message="Click below to continue to the <strong>📈 Maturity Assessment</strong> "
                    "with your DQ score pre-filled.",
            color="#a78bfa",
        )

        st.divider()
        col1, col2, col3 = st.columns([1, 1, 1])
        with col2:
            if st.button("📈 Continue to Maturity Assessment →",
                         type="primary", use_container_width=True, key="dq_to_mat"):
                st.session_state["page"] = "maturity"; st.rerun()

    except Exception as e:
        st.markdown('<div class="banner danger">', unsafe_allow_html=True)
        st.markdown(f"❌ **Error:** {e}")
        st.markdown('</div>', unsafe_allow_html=True)
        with st.expander("🔍 Technical Details"):
            st.code(traceback.format_exc())


# ══════════════════════════════════════════════════════════════════════════
#  PAGE: MATURITY ASSESSMENT
# ══════════════════════════════════════════════════════════════════════════
def _apply_editor_edits(dim: str, editor_key: str) -> None:
    widget_state = st.session_state.get(editor_key)
    if not widget_state:
        return
    edited_rows = widget_state.get("edited_rows", {})
    if not edited_rows:
        return
    df = st.session_state.mat_responses[dim].copy()
    for row_idx, changes in edited_rows.items():
        for col, val in changes.items():
            df.at[int(row_idx), col] = val
    st.session_state.mat_responses[dim] = df


def _do_submit() -> None:
    objects   = st.session_state.mat_objects
    dims      = st.session_state.mat_dims
    responses = st.session_state.mat_responses
    cn        = st.session_state.mat_client_name or "Client"
    bm        = float(st.session_state.mat_benchmark)
    tg        = float(st.session_state.mat_target)
    lt        = float(st.session_state.mat_low_thr)
    dq_score  = st.session_state.get("dq_score")

    ok, msg = validate_responses(responses, dims, objects)
    if not ok:
        st.error(f"⚠️ Validation failed: {msg}")
        return

    with st.spinner("⚙️ Computing scores and building reports…"):
        dim_table, overall = compute_all_scores(objects, dims, responses)
        domain_display = {
            dim: float(np.nanmean(dim_table.loc[dim].values)) for dim in dims
        }
        exec_score = float(np.nanmean(overall.values)) if len(overall) else 0.0
        slide_png  = render_slide_png(
            client_name=cn, domain_scores=domain_display,
            exec_score=exec_score if np.isfinite(exec_score) else 0.0,
            benchmark=bm, target=tg,
        )
        pdf_bytes = build_pdf_bytes(
            client_name=cn, slide_png=slide_png, dim_table=dim_table,
            overall=overall, detail_tables=responses, dq_score=dq_score,
        )
        mat_excel = to_excel_bytes(
            dim_table=dim_table, overall=overall, detail_tables=responses,
            low_thr=lt, objects=objects,
        )
        combined_excel = (
            _combined_excel(dq_score, st.session_state.get("dq_dim_scores"), mat_excel)
            if dq_score is not None else mat_excel
        )

    ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    st.session_state["mat_submitted"] = True
    st.session_state["mat_payload"]   = {
        "dim_table": dim_table, "overall": overall,
        "slide_png": slide_png, "mat_excel": mat_excel,
        "combined_excel": combined_excel, "pdf_bytes": pdf_bytes,
        "client_name": cn, "ts": ts,
    }
    st.rerun()


def page_maturity():
    inject_gdg_light()
    dq_score  = st.session_state.get("dq_score")
    submitted = st.session_state.get("mat_submitted", False)

    with st.sidebar:
        st.markdown("### 🧭 Navigation")
        if st.button("🏠 Home",     use_container_width=True, key="mat_home"):
            st.session_state["page"] = "home"; st.rerun()
        if st.button("🔍 DQ",       use_container_width=True, key="mat_dq"):
            st.session_state["page"] = "dq"; st.rerun()
        if st.button("📋 Policies", use_container_width=True, key="mat_policy"):
            st.session_state["page"] = "policy"; st.rerun()
        if st.button("🎯 Cases",    use_container_width=True, key="mat_cases"):
            st.session_state["page"] = "cases"; st.rerun()
        st.divider()

        st.markdown("### ⚙️ Configuration")
        st.session_state["mat_client_name"] = st.text_input(
            "Client Name", value=st.session_state.get("mat_client_name", ""),
            placeholder="Organisation name", disabled=submitted,
        )
        all_obj_opts = list(dict.fromkeys(
            DEFAULT_MASTER_OBJECTS + st.session_state.mat_objects
        ))
        st.session_state["mat_objects"] = st.multiselect(
            "Master Data Objects", options=all_obj_opts,
            default=st.session_state.mat_objects, disabled=submitted,
        )
        st.session_state["mat_dims"] = st.multiselect(
            "Maturity Dimensions", options=MATURITY_DIMS,
            default=st.session_state.mat_dims, disabled=submitted,
        )
        st.divider()
        st.markdown("### 📊 Thresholds")
        st.session_state["mat_low_thr"] = st.slider(
            "Exception threshold (≤)", 1.0, 5.0,
            float(st.session_state.get("mat_low_thr", 2.0)), 0.5, disabled=submitted,
        )
        st.divider()
        st.markdown("### 🎯 Benchmark / Target")
        st.session_state["mat_benchmark"] = st.number_input(
            "Industry Benchmark", 1.0, 5.0,
            float(st.session_state.get("mat_benchmark", 3.0)), 0.1, disabled=submitted,
        )
        st.session_state["mat_target"] = st.number_input(
            "Target Score", 1.0, 5.0,
            float(st.session_state.get("mat_target", 3.0)), 0.1, disabled=submitted,
        )

    if not st.session_state.mat_objects or not st.session_state.mat_dims:
        st.markdown('<div class="info-box">', unsafe_allow_html=True)
        st.markdown("👉 Please select at least one **Object** and one **Dimension** in the sidebar.")
        st.markdown('</div>', unsafe_allow_html=True)
        st.stop()

    # ── DEFERRED SYNC FIX ──────────────────────────────────────────────────
    # When the user changes objects/dims in the multiselect, Streamlit reruns
    # immediately. If we call sync_response_tables() in that same rerun, it
    # rebuilds DataFrames and pops editor snapshot keys while Streamlit is
    # still reconciling widget state — causing the dropdown to lose focus and
    # discard the selection (requiring multiple clicks).
    #
    # Fix: on the first rerun after a change, just set a _sync_pending flag
    # and trigger another rerun. Only on the second rerun (when widget state
    # is fully settled) do we actually perform the sync.
    # ───────────────────────────────────────────────────────────────────────
    prev_objs = st.session_state.get("_last_sync_objects")
    prev_dims = st.session_state.get("_last_sync_dims")
    curr_objs = st.session_state.mat_objects
    curr_dims = st.session_state.mat_dims

    needs_sync = (
        prev_objs is None
        or prev_dims is None
        or set(prev_objs) != set(curr_objs)
        or set(prev_dims) != set(curr_dims)
    )

    if needs_sync:
        if st.session_state.get("_sync_pending"):
            # Second rerun: widget state is settled, safe to sync now
            sync_response_tables()
            for d in curr_dims:
                st.session_state.pop(f"mat_snap_{d}", None)
            st.session_state["_last_sync_objects"] = list(curr_objs)
            st.session_state["_last_sync_dims"]    = list(curr_dims)
            st.session_state["_sync_pending"] = False
        else:
            # First rerun after change: just flag and rerun again
            st.session_state["_sync_pending"] = True
            st.rerun()

    if dq_score is not None and not st.session_state.get("dq_autofilled"):
        autofill_dq_dimension(dq_score)
        st.session_state.pop("mat_snap_Data Quality", None)
        st.session_state["dq_autofilled"] = True

    # ── REPORT VIEW ────────────────────────────────────────────────────────
    if submitted and st.session_state.get("mat_payload"):
        p  = st.session_state["mat_payload"]
        cn = p["client_name"]
        ts = p["ts"]

        st.markdown("# ✅ Data Maturity Assessment Report")

        if dq_score is not None:
            lvl = dq_score_to_maturity_level(dq_score)
            st.markdown(
                f'<div class="banner success">'
                f'**DQ Engine Score:** {dq_score:.1f}% → **Level:** {lvl} (applied to Data Quality dimension)'
                f'</div>',
                unsafe_allow_html=True,
            )

        st.markdown("### 📊 Summary Slide")
        st.image(p["slide_png"], use_container_width=True)

        UIComponents.render_micro_progress(100, "#10b981", "#34d399")
        st.divider()

        st.markdown("""
            <style>
            .dataframe tbody tr:hover { background-color: rgba(224,242,254,0.5) !important; }
            .dataframe thead th {
                background: linear-gradient(135deg,#e0f2fe 0%,#bae6fd 100%) !important;
                color: #0c4a6e !important;
            }
            </style>
        """, unsafe_allow_html=True)

        t1, t2 = st.columns(2)
        with t1:
            st.markdown("#### Dimension-wise Maturity")
            styled_dim = p["dim_table"].style\
                .format("{:.2f}")\
                .background_gradient(cmap="Blues", axis=None, vmin=1, vmax=5)
            st.dataframe(styled_dim, use_container_width=True)
        with t2:
            st.markdown("#### Overall Maturity Score")
            styled_overall = pd.DataFrame(p["overall"]).T.style\
                .format("{:.2f}")\
                .background_gradient(cmap="Blues", axis=None, vmin=1, vmax=5)
            st.dataframe(styled_overall, use_container_width=True)

        st.divider()
        st.markdown("#### Scores by Dimension")
        dim_vals = {
            dim: float(np.nanmean(p["dim_table"].loc[dim].values))
            for dim in p["dim_table"].index
        }
        bar_img = _mat_bar_png(dim_vals)
        if bar_img:
            st.image(bar_img, use_container_width=True)

        st.divider()
        st.markdown("### 📥 Download Reports")
        safe_cn = cn.replace(" ", "_")

        UIComponents.render_hint_chip(
            "3 formats available", tip="PDF · Maturity Excel · Combined DQ+Maturity Excel", icon="📥"
        )
        st.markdown('<div style="height:0.5rem;"></div>', unsafe_allow_html=True)

        d1, d2, d3 = st.columns(3)
        with d1:
            pdf_filename = get_timestamp_filename(f"Maturity_Report_{safe_cn}", "pdf")
            st.download_button(
                "📄 PDF Report", data=p["pdf_bytes"],
                file_name=pdf_filename, mime="application/pdf",
                use_container_width=True,
            )
        with d2:
            mat_excel_filename = get_timestamp_filename(f"Maturity_Assessment_{safe_cn}", "xlsx")
            st.download_button(
                "📊 Maturity Excel", data=p["mat_excel"],
                file_name=mat_excel_filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        with d3:
            combined_filename = get_timestamp_filename(f"DQ_Maturity_Combined_{safe_cn}", "xlsx")
            st.download_button(
                "🔗 Combined Excel", data=p["combined_excel"],
                file_name=combined_filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )

        st.divider()
        if st.button("✏️ Edit Responses", use_container_width=True, key="mat_edit"):
            st.session_state["mat_submitted"] = False
            st.session_state["mat_payload"]   = {}
            st.rerun()
        st.stop()

    # ── QUESTIONNAIRE VIEW ─────────────────────────────────────────────────
    st.markdown("# 📈 Data Maturity Assessment")

    if dq_score is not None:
        lvl = dq_score_to_maturity_level(dq_score)
        st.markdown(
            f'<div class="banner success">'
            f'✅ DQ Score **{dq_score:.1f}%** → level **{lvl}** auto-filled in *Data Quality* dimension.'
            f'</div>',
            unsafe_allow_html=True,
        )

    UIComponents.render_action_hint_bar(
        title="How to complete",
        message="Select a rating for each question using the "
                "<strong>dropdown columns</strong>. Weighted scoring is automatic.",
        color="#a78bfa",
    )
    st.divider()

    dims = st.session_state.mat_dims
    tabs = st.tabs(dims)

    for i, dim in enumerate(dims):
        with tabs[i]:
            st.markdown(f"### {dim}")
            if dim == "Data Quality" and dq_score is not None:
                lvl = dq_score_to_maturity_level(dq_score)
                st.markdown(
                    f'<div class="banner">'
                    f'Auto-populated from DQ Score **{dq_score:.1f}%** → **{lvl}**. '
                    f'You can adjust individual ratings as needed.'
                    f'</div>',
                    unsafe_allow_html=True,
                )

            cfg = {"Weight": st.column_config.NumberColumn("Weight", min_value=0.0, step=0.5)}
            for obj in st.session_state.mat_objects:
                cfg[obj] = st.column_config.SelectboxColumn(obj, options=RATING_LABELS, required=True)

            editor_key = f"mat_editor_{dim}"
            st.data_editor(
                st.session_state.mat_responses[dim],
                use_container_width=True, hide_index=True,
                column_config=cfg,
                disabled=["Question ID", "Section", "Question"],
                key=editor_key,
                on_change=_apply_editor_edits,
                args=(dim, editor_key),
            )

    st.divider()
    col1, col2, col3, col4 = st.columns([1, 1, 1, 1])
    with col2:
        if st.button("🚀 Submit & Generate Report", type="primary",
                     use_container_width=True, key="mat_submit"):
            _do_submit()
    with col3:
        st.info("**Submit** to generate visuals and downloadable reports.")


# ══════════════════════════════════════════════════════════════════════════
#  PAGE: POLICY HUB
# ══════════════════════════════════════════════════════════════════════════
def page_policy_hub():
    with st.sidebar:
        st.markdown("### 🧭 Navigation")
        if st.button("🏠 Home",      use_container_width=True, key="policy_home"):
            st.session_state["page"] = "home"; st.rerun()
        if st.button("🔍 DQ",        use_container_width=True, key="policy_dq"):
            st.session_state["page"] = "dq"; st.rerun()
        if st.button("📈 Maturity",  use_container_width=True, key="policy_maturity"):
            st.session_state["page"] = "maturity"; st.rerun()
        if st.button("🎯 Cases",     use_container_width=True, key="policy_cases"):
            st.session_state["page"] = "cases"; st.rerun()

    st.markdown("""
        <div class="policy-hero">
            <h1>📋 Policy Hub & Procedures Management</h1>
            <p>Centralized repository for enterprise data governance policies, procedures, and standards</p>
        </div>
    """, unsafe_allow_html=True)

    UIComponents.render_action_hint_bar(
        title="Browse Modules",
        message="Expand each card below to explore the <strong>Policy Hub</strong> feature set.",
        color="#c084fc",
    )

    st.markdown("""
    <div class="ph-section-intro">
        <p>
        The <strong>Policy Hub by Uniqus</strong> is a centralized platform that helps organizations manage
        policies, procedures, and approvals in one place. It offers an easy-to-use interface where users can
        upload documents, track workflows, receive notifications, and ensure compliance.
        Browse the modules below to see what users can do within the frontend system.
        </p>
    </div>
    """, unsafe_allow_html=True)

    with st.expander("📁  Workflow Automation  ✅", expanded=False):
        st.markdown("""
        <div class="ph-feature-item">
            <div class="ph-feature-icon">🚀</div>
            <div class="ph-feature-content">
                <div class="ph-feature-title">Submit for Approval Button</div>
                <div class="ph-feature-desc">Users can send a policy to reviewers in one click.</div>
            </div>
        </div>
        <div class="ph-feature-item">
            <div class="ph-feature-icon">📊</div>
            <div class="ph-feature-content">
                <div class="ph-feature-title">Workflow Status Tracker</div>
                <div class="ph-feature-desc">Shows current stage of the policy lifecycle:</div>
                <div class="ph-sub-list">
                    <span class="ph-sub-tag">Draft</span>
                    <span class="ph-sub-tag">Under Review</span>
                    <span class="ph-sub-tag">Approved</span>
                    <span class="ph-sub-tag">Published</span>
                </div>
            </div>
        </div>
        <div class="ph-feature-item">
            <div class="ph-feature-icon">🕐</div>
            <div class="ph-feature-content">
                <div class="ph-feature-title">Approval Timeline View</div>
                <div class="ph-feature-desc">Displays who approved, rejected, or reviewed — and when.</div>
            </div>
        </div>
        <div class="ph-feature-item">
            <div class="ph-feature-icon">📋</div>
            <div class="ph-feature-content">
                <div class="ph-feature-title">Pending Actions Panel</div>
                <div class="ph-feature-desc">Users can see tasks waiting for their approval at a glance.</div>
            </div>
        </div>
        <div class="ph-feature-item">
            <div class="ph-feature-icon">📧</div>
            <div class="ph-feature-content">
                <div class="ph-feature-title">Email Approval Links</div>
                <div class="ph-feature-desc">Approvers can approve or reject directly from email — no need to log in.</div>
            </div>
        </div>
        <div class="ph-feature-item">
            <div class="ph-feature-icon">⚠️</div>
            <div class="ph-feature-content">
                <div class="ph-feature-title">Escalation Alerts</div>
                <div class="ph-feature-desc">If approval is delayed, the system highlights it and sends escalation notifications.</div>
            </div>
        </div>
        <div class="ph-benefit-banner">
            <div class="ph-bb-icon">💡</div>
            <div class="ph-bb-content">
                <div class="ph-bb-label">User Benefit</div>
                <div class="ph-bb-text">No manual tracking — everything is automated and visible.</div>
            </div>
        </div>
        """, unsafe_allow_html=True)

    with st.expander("📁  Notification & Reminders  ✅", expanded=False):
        st.markdown("""
        <div class="ph-feature-item">
            <div class="ph-feature-icon">🔔</div>
            <div class="ph-feature-content">
                <div class="ph-feature-title">Notification Bell Icon</div>
                <div class="ph-feature-desc">Shows real-time alerts inside the portal with unread count badge.</div>
            </div>
        </div>
        <div class="ph-feature-item">
            <div class="ph-feature-icon">📬</div>
            <div class="ph-feature-content">
                <div class="ph-feature-title">In-App Notification List</div>
                <div class="ph-feature-desc">Displays actionable messages in real time:</div>
                <div class="ph-sub-list">
                    <span class="ph-sub-tag">Policy approved</span>
                    <span class="ph-sub-tag">Review requested</span>
                    <span class="ph-sub-tag">Comments added</span>
                </div>
            </div>
        </div>
        <div class="ph-feature-item">
            <div class="ph-feature-icon">📧</div>
            <div class="ph-feature-content">
                <div class="ph-feature-title">Email Notifications</div>
                <div class="ph-feature-desc">Users receive alerts directly in Outlook / email for all policy events.</div>
            </div>
        </div>
        <div class="ph-feature-item">
            <div class="ph-feature-icon">⏰</div>
            <div class="ph-feature-content">
                <div class="ph-feature-title">Reminder Alerts</div>
                <div class="ph-feature-desc">Proactive notifications for upcoming and overdue items:</div>
                <div class="ph-sub-list">
                    <span class="ph-sub-tag">Pending approvals</span>
                    <span class="ph-sub-tag">Overdue tasks</span>
                    <span class="ph-sub-tag">Policy review due dates</span>
                </div>
            </div>
        </div>
        <div class="ph-feature-item">
            <div class="ph-feature-icon">⚙️</div>
            <div class="ph-feature-content">
                <div class="ph-feature-title">Digest Settings</div>
                <div class="ph-feature-desc">Users can choose their preferred notification frequency:</div>
                <div class="ph-sub-list">
                    <span class="ph-sub-tag">Instant alerts</span>
                    <span class="ph-sub-tag">Daily summary</span>
                    <span class="ph-sub-tag">Weekly summary</span>
                </div>
            </div>
        </div>
        <div class="ph-benefit-banner">
            <div class="ph-bb-icon">💡</div>
            <div class="ph-bb-content">
                <div class="ph-bb-label">User Benefit</div>
                <div class="ph-bb-text">Users never miss approvals or deadlines.</div>
            </div>
        </div>
        """, unsafe_allow_html=True)

    with st.expander("📁  Role-Based User Access  ✅", expanded=False):
        st.markdown("""
        <div class="ph-feature-item">
            <div class="ph-feature-icon">🏠</div>
            <div class="ph-feature-content">
                <div class="ph-feature-title">Role-Based Dashboard</div>
                <div class="ph-feature-desc">Different homepages tailored for each user role:</div>
                <div class="ph-sub-list">
                    <span class="ph-sub-tag">Admin</span>
                    <span class="ph-sub-tag">Editor</span>
                    <span class="ph-sub-tag">Reviewer</span>
                    <span class="ph-sub-tag">Viewer</span>
                </div>
            </div>
        </div>
        <div class="ph-feature-item">
            <div class="ph-feature-icon">🔒</div>
            <div class="ph-feature-content">
                <div class="ph-feature-title">Restricted Document View</div>
                <div class="ph-feature-desc">Sensitive policies are visible only to authorized users based on their clearance level.</div>
            </div>
        </div>
        <div class="ph-feature-item">
            <div class="ph-feature-icon">🎛️</div>
            <div class="ph-feature-content">
                <div class="ph-feature-title">Edit / View Controls</div>
                <div class="ph-feature-desc">Buttons like Edit, Publish, Delete appear only if the user has the required permission.</div>
            </div>
        </div>
        <div class="ph-feature-item">
            <div class="ph-feature-icon">🏢</div>
            <div class="ph-feature-content">
                <div class="ph-feature-title">Department Filtering</div>
                <div class="ph-feature-desc">Users see policies related to their own department automatically.</div>
            </div>
        </div>
        <div class="ph-feature-item">
            <div class="ph-feature-icon">🔑</div>
            <div class="ph-feature-content">
                <div class="ph-feature-title">Secure Login (SSO)</div>
                <div class="ph-feature-desc">Login seamlessly using company credentials via Azure AD Single Sign-On.</div>
            </div>
        </div>
        <div class="ph-benefit-banner">
            <div class="ph-bb-icon">💡</div>
            <div class="ph-bb-content">
                <div class="ph-bb-label">User Benefit</div>
                <div class="ph-bb-text">Ensures security while keeping the UI simple and clutter-free.</div>
            </div>
        </div>
        """, unsafe_allow_html=True)

    with st.expander("📁  White-Labelling of Tool  ✅", expanded=False):
        st.markdown("""
        <div class="ph-feature-item">
            <div class="ph-feature-icon">🎨</div>
            <div class="ph-feature-content">
                <div class="ph-feature-title">Company Logo & Branding</div>
                <div class="ph-feature-desc">Portal displays company logo, corporate colors, and approved fonts throughout.</div>
            </div>
        </div>
        <div class="ph-feature-item">
            <div class="ph-feature-icon">🏠</div>
            <div class="ph-feature-content">
                <div class="ph-feature-title">Custom Homepage Layout</div>
                <div class="ph-feature-desc">Dashboard designed as per specific business needs and organizational structure.</div>
            </div>
        </div>
        <div class="ph-feature-item">
            <div class="ph-feature-icon">📧</div>
            <div class="ph-feature-content">
                <div class="ph-feature-title">Branded Email Templates</div>
                <div class="ph-feature-desc">Approval and notification emails follow company branding guidelines.</div>
            </div>
        </div>
        <div class="ph-feature-item">
            <div class="ph-feature-icon">🌗</div>
            <div class="ph-feature-content">
                <div class="ph-feature-title">Theme Options</div>
                <div class="ph-feature-desc">Light / Dark mode selection for comfortable viewing experience.</div>
            </div>
        </div>
        <div class="ph-feature-item">
            <div class="ph-feature-icon">🧩</div>
            <div class="ph-feature-content">
                <div class="ph-feature-title">Personalized Widgets</div>
                <div class="ph-feature-desc">Users can add or remove dashboard widgets as needed:</div>
                <div class="ph-sub-list">
                    <span class="ph-sub-tag">My Tasks</span>
                    <span class="ph-sub-tag">Recent Policies</span>
                    <span class="ph-sub-tag">Pending Approvals</span>
                </div>
            </div>
        </div>
        <div class="ph-benefit-banner">
            <div class="ph-bb-icon">💡</div>
            <div class="ph-bb-content">
                <div class="ph-bb-label">User Benefit</div>
                <div class="ph-bb-text">The tool feels like your own — fully branded, familiar, and trusted.</div>
            </div>
        </div>
        """, unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════
#  START APPLICATION
# ══════════════════════════════════════════════════════════════════════════
load_css()
_init_state()

{
    "home":     page_home,
    "dq":       page_dq,
    "maturity": page_maturity,
    "policy":   page_policy_hub,
    "cases":    page_case_management,
}[st.session_state.page]()